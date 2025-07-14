from datetime import datetime
from .twilio_whatsapp import send_registration_success_whatsapp
import openpyxl # Import openpyxl for Excel generation
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .forms import HalaqaForm, ExcelUploadForm # Import ExcelUploadForm
from .forms import StudentForm, TeacherForm, LoginForm, DailyMemorizationTaskForm, DailyReviewTaskForm, DailyTalqeenTaskForm, WeeklyLectureForm, WeeklyPlanForm, DepartmentForm, MsarForm, HalaqaForm, StudentLinkForm
from .models import Students, Teachers, Halaqa, UserRole, Supervisor, User, DailyMemorizationTask, DailyReviewTask, DailyTalqeenTask, WeeklyLecture, WeeklyPlan, Department, Msar, Language, StudentLink
from django.http import HttpResponse, HttpResponseForbidden, JsonResponse, HttpResponseNotFound, Http404
from django.contrib.auth.models import Group, User
from django.core import serializers
from django.db.models import F, Sum, Q, IntegerField # Import Sum, Q, and IntegerField
from django.db.models.functions import Coalesce # Import Coalesce
from datetime import date # Import date
from django.views.decorators.csrf import csrf_exempt # Import csrf_exempt
import json
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
from .whatsapp_notifier import send_whatsapp_message
from django.db import transaction # Import transaction for atomic operations
from django.contrib.auth.hashers import make_password # Import make_password




# In views.py

def handle_form_submission(request, form_class, template_name, success_url, success_message):
    if request.method == 'POST':
        form = form_class(request.POST, request.FILES or None)
        print(request.POST)
        print(form.is_valid())
        if form.is_valid():
            form.save()
            messages.success(request, success_message)
            return redirect(success_url)
        else:
            print(form.errors)
            messages.error(request, "There was an error in the form. Please check your input.")
    else:
        form = form_class()
    return render(request, template_name, {'form': form})



def teacher_register(request):
    if request.method == 'POST':
        print("Request FILES:", request.FILES) # Add this line
        if 'cv_teacher' in request.FILES:
            cv_file = request.FILES['cv_teacher']
            print(f"CV File Name: {cv_file.name}")
            print(f"CV File Size: {cv_file.size}")
            print(f"CV File Content Type: {cv_file.content_type}")
        form = TeacherForm(request.POST, request.FILES)
        if form.is_valid():
            username = form.cleaned_data['username']
            # Check if username already exists
            if User.objects.filter(username=username).exists():
                form.add_error('username', 'اسم المستخدم موجود بالفعل. يرجى اختيار اسم مستخدم آخر.')
                messages.error(request, "حدث خطأ أثناء التسجيل. اسم المستخدم موجود بالفعل.")
                return render(request, 'teacher/register_teacher.html', {'form': form, 'errors': form.errors})

            # Create a Django User object
            user = User.objects.create_user(
                username=username,
                password=form.cleaned_data['password'],
                email=form.cleaned_data.get('email') or None, # Use .get() for optional email, pass None if empty
                first_name=form.cleaned_data['first_name'],
                last_name=form.cleaned_data['last_name']
            )
            user.save()

            # Create the Teacher instance and link it to the created User
            teacher = form.save(commit=False)
            teacher.user = user
            teacher.save() # Save the teacher instance to create the primary key

            # Handle multiple language selection and set default to Arabic if none selected
            selected_languages = form.cleaned_data.get('language')
            if not selected_languages:
                # If no languages are selected, default to Arabic
                arabic_language, created = Language.objects.get_or_create(name='Arabic')
                teacher.language.set([arabic_language])
            else:
                # Set the selected languages
                teacher.language.set([selected_languages])


            # Add the user to the "teachers" group
            print("Attempting to add user to 'teachers' group.")
            try:
                teachers_group, created = Group.objects.get_or_create(name='Teachers')
                if created:
                    print("Created 'Teachers' group.")
                else:
                    print("Found existing 'Teachers' group.")
                user.groups.add(teachers_group)
                print(f"User {user.username} added to 'Teachers' group successfully.")
            except Exception as e:
                print(f"Error adding user to 'Teachers' group: {e}")


            # Send WhatsApp notification
            whatsapp_number = teacher.mobile_whatsapp
            print(f"Attempting to send WhatsApp to: {whatsapp_number}") # Debug print
            if whatsapp_number:
                print("WhatsApp number found, attempting to send message.") # Debug print
                send_registration_success_whatsapp(str(whatsapp_number)) # Convert PhoneNumberField to string
            else:
                print("No WhatsApp number found for the teacher.") # Debug print

            messages.success(request, "تم التسجيل بنجاح!")
            messages.success(request, "تم التسجيل بنجاح!")
            print("Redirecting to teacher_registration_success.") # Debug print
            # Redirect to the teacher registration success page
            return redirect('teacher_registration_success')
        else:
            messages.error(request, "حدث خطأ أثناء التسجيل. يرجى مراجعة الأخطاء أدناه.")
            # Pass form errors to the template
            return render(request, 'teacher/register_teacher.html', {'form': form, 'errors': form.errors})

    else:
        form = TeacherForm()

    return render(request, 'teacher/register_teacher.html', {'form': form})
def teacher_registration_success(request):
    return render(request, 'teacher/teacher_registration_success.html')






from django.contrib.auth.models import Group
from django.contrib import messages
from django.shortcuts import redirect, render
from .models import Students, TypeRead
    
from .forms import StudentForm, TypeReadForm
from django.contrib.auth.models import User

def student_register(request):
    if request.method == 'POST':
        print("request.POST:", request.POST)
        print("request.POST.getlist('mobile'):", request.POST.getlist('mobile')) # Check list format
        print("request.POST.getlist('whatsapp'):", request.POST.getlist('whatsapp')) # Check list format
        form = StudentForm(request.POST)
        if form.is_valid():
            user_data = form.cleaned_data
            if User.objects.filter(username=user_data['username']).exists():
                messages.error(request, "اسم المستخدم موجود بالفعل. يرجى اختيار اسم مستخدم آخر.")
            else:
                user = User.objects.create_user(
                    username=user_data['username'],
                    password=user_data['password'],
                    email=user_data['email'],
                    first_name=user_data['first_name'],
                    last_name=user_data['last_name']
                )
                user.save()

                # Create UserRole for the new student
                user_role, created_role = UserRole.objects.get_or_create(user=user)
                user_role.role = 'student'
                user_role.save()
                print(f"User {user.username} assigned role 'student'.") # Debug print


                student = Students(
                    user=user,
                    first_name=user_data['first_name'],
                    second_name=user_data['second_name'],
                    last_name=user_data['last_name'],
                    birthday=user_data['birthday'],
                    nationality=user_data['nationality'],
                    gender=user_data['gender'],
                    email=user_data['email'],
                    language=user_data['language'],
                    mobile=user_data['mobile'],
                    whatsapp=user_data['whatsapp'],
                    username=user_data['username'],
                    password=user_data['password']
                )
                student.save()

                students_group, created = Group.objects.get_or_create(name='students')
                user.groups.add(students_group)

                # Send WhatsApp notification
                whatsapp_number = student.whatsapp
                print(f"Attempting to send WhatsApp to student: {whatsapp_number}") # Debug print
                if whatsapp_number:
                    print("Student WhatsApp number found, attempting to send message.") # Debug print
                    send_registration_success_whatsapp(str(whatsapp_number)) # Convert PhoneNumberField to string
                else:
                    print("No WhatsApp number found for the student.") # Debug print

                messages.success(request, "تم إضافة الطالب بنجاح!")
                return render(request, 'students/registration_success.html', {'user': user})
        else:
            messages.error(request, "حدث خطأ أثناء التسجيل.")
    else:
        form = StudentForm()

    return render(request, 'students/register_student.html', {'form': form})




from django.shortcuts import render

def registration_success(request, user):
    return render(request, 'students/registration_success.html', {'user': user})






def custom_login_view(request):
    if request.method == 'POST':
        form = LoginForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                try:
                    user_role = UserRole.objects.get(user=user)
                    if user_role.role == 'student':
                        student = Students.objects.get(username=username)
                        return redirect('student_home', student_id=student.id)
                    elif user_role.role == 'teacher':
                        teacher = Teachers.objects.get(user=user) # Get teacher object using the user
                        # Send WhatsApp message to the teacher's number
                        teacher_whatsapp_number = teacher.mobile_whatsapp
                        if teacher_whatsapp_number:
                            send_registration_success_whatsapp(str(teacher_whatsapp_number)) # Call the function
                        return redirect('teacher_home', username=username)
                    elif user_role.role == 'supervisor':
                        return redirect('supervisor_dashboard')
                    elif user_role.role == 'admin':
                        return redirect('admin_dashboard')
                    else:
                        return redirect('index')
                except UserRole.DoesNotExist:
                    messages.error(request, "Role does not exist for this user.")
                    return redirect('login')
                except Teachers.DoesNotExist:
                     messages.error(request, "Teacher profile not found for this user.")
                     return redirect('login')
            else:
                messages.error(request, "Invalid username or password.")
                return redirect('login')
    else:
        form = LoginForm()
    return render(request, 'login/login.html', {'form': form})








@login_required
def teacher_home(request, username):
    # الحصول على كائن المعلم بناءً على اسم المستخدم
    teacher = get_object_or_404(Teachers, user__username=username)
    classes = Halaqa.objects.filter(teacher=teacher)
    return render(request, 'teacher/teacher_home.html', {'teacher': teacher, 'classes': classes})






@login_required
def student_home_view(request, student_id):
    student = get_object_or_404(Students, id=student_id)
    lecture = WeeklyLecture.objects.order_by('-created_at').first()
    student_links = student.links.all() # Fetch all links for the student
    context = {
        'student': student,
        'lecture': lecture,
        'student_links': student_links, # Add student_links to the context
    }
    return render(request, 'students/student_home.html', context)






@login_required
def superuser_dashboard(request):
    return render(request, 'teacher/superuser_dashboard.html')

def admin_dashboard(request):
    return render(request, 'teacher/admin_dashboard.html')

@login_required
def manage_weekly_lectures_view(request):
    # Check if the user is an admin
    if not request.user.groups.filter(name='admin').exists():
        messages.error(request, "You do not have permission to access this page.")
        return redirect('index') # Or a suitable unauthorized page

    weekly_lectures = WeeklyLecture.objects.all().order_by('-created_at')

    context = {
        'weekly_lectures': weekly_lectures
    }
    return render(request, 'admin/manage_weekly_lectures.html', context)


@login_required
def add_weekly_lecture_view(request):
    # Check if the user is an admin
    if not request.user.groups.filter(name='admin').exists():
        messages.error(request, "You do not have permission to access this page.")
        return redirect('index') # Or a suitable unauthorized page

    if request.method == 'POST':
        form = WeeklyLectureForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "تم إضافة المحاضرة بنجاح!")
            return redirect('manage_weekly_lectures') # Redirect to the management page after adding
        else:
            messages.error(request, "حدث خطأ أثناء إضافة المحاضرة. يرجى مراجعة الأخطاء أدناه.")
    else:
        form = WeeklyLectureForm()

    return render(request, 'admin/add_weekly_lecture.html', {'form': form})


@login_required
def edit_weekly_lecture_view(request, lecture_id):
    # Check if the user is an admin
    if not request.user.groups.filter(name='admin').exists():
        messages.error(request, "You do not have permission to access this page.")
        return redirect('index') # Or a suitable unauthorized page

    lecture = get_object_or_404(WeeklyLecture, id=lecture_id)

    if request.method == 'POST':
        form = WeeklyLectureForm(request.POST, instance=lecture)
        if form.is_valid():
            form.save()
            messages.success(request, "تم تحديث المحاضرة بنجاح!")
            return redirect('manage_weekly_lectures') # Redirect to the management page after editing
        else:
            messages.error(request, "حدث خطأ أثناء تحديث المحاضرة. يرجى مراجعة الأخطاء أدناه.")
    else:
        form = WeeklyLectureForm(instance=lecture)

    return render(request, 'admin/edit_weekly_lecture.html', {'form': form, 'lecture': lecture})


@login_required
def statistics_dashboard_view(request):
    # Check if the user is an admin
    if not request.user.groups.filter(name='admin').exists():
        messages.error(request, "You do not have permission to access this page.")
        return redirect('index') # Or a suitable unauthorized page

    # Get data for Masarat chart (count of students per Msar)
    masarat_stats = Msar.objects.annotate(student_count=Sum('halaqa__students')).order_by('name')

    # Get data for Halaqat chart (count of students per Halaqa)
    halaqat_stats = Halaqa.objects.annotate(student_count=Sum('students')).order_by('name')

    # Get data for Hifz chart (total memorized pages per Msar)
    hifz_stats_by_msar = Msar.objects.annotate(
        total_pages_and_parts=Coalesce(Sum('halaqa__students__dailymemorizationtask__pages', output_field=IntegerField()), 0) +
                              Coalesce(Sum('halaqa__students__dailymemorizationtask__partial_page'), 0.0)
    ).order_by('name')

    # Prepare data for Masarat chart
    masarat_chart_data = []
    for msar in masarat_stats:
        masarat_chart_data.append({
            'name': msar.name,
            'student_count': float(msar.student_count) if msar.student_count is not None else 0.0
        })

    # Prepare data for Halaqat chart
    halaqat_chart_data = []
    for halaqa in halaqat_stats:
        halaqat_chart_data.append({
            'name': halaqa.name,
            'student_count': float(halaqa.student_count) if halaqa.student_count is not None else 0.0
        })

    # Prepare data for Hifz chart (by Msar)
    hifz_chart_data = []
    for msar in hifz_stats_by_msar:
        hifz_chart_data.append({
            'name': msar.name,
            'total_pages_and_parts': float(msar.total_pages_and_parts) if msar.total_pages_and_parts is not None else 0.0
        })

    context = {
        'masarat_stats_json': json.dumps(masarat_chart_data),
        'halaqat_stats_json': json.dumps(halaqat_chart_data),
        'hifz_stats_by_msar_json': json.dumps(hifz_chart_data),
    }
    return render(request, 'admin/statistics_dashboard.html', context)

@login_required
def manage_report_view(request):
    # Check if the user is an admin
    if not request.user.groups.filter(name='admin').exists():
        messages.error(request, "You do not have permission to access this page.")
        return redirect('index') # Or a suitable unauthorized page

    # Add a cache-busting parameter (e.g., timestamp)
    cache_buster = int(time.time())

    context = {
        'cache_buster': cache_buster
    }

    return render(request, 'admin/manage_report.html', context)

@login_required
def halaqa_report_view(request):
    # Check if the user is an admin
    if not request.user.groups.filter(name='admin').exists():
        messages.error(request, "You do not have permission to access this page.")
        return redirect('index') # Or a suitable unauthorized page

    halaqat = Halaqa.objects.all()
    selected_halaqa_id = request.GET.get('halaqa_id')
    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')

    selected_halaqa = None
    memorization_records = None
    review_records = None
    talqeen_records = None
    from_date = None
    to_date = None

    if selected_halaqa_id:
        try:
            selected_halaqa = Halaqa.objects.get(id=selected_halaqa_id)
            students_in_halaqa = Students.objects.filter(halaqa=selected_halaqa)

            memorization_records = DailyMemorizationTask.objects.filter(student__in=students_in_halaqa)
            review_records = DailyReviewTask.objects.filter(student__in=students_in_halaqa)
            talqeen_records = DailyTalqeenTask.objects.filter(student__in=students_in_halaqa)

            if from_date_str and to_date_str:
                try:
                    from_date = date.fromisoformat(from_date_str)
                    to_date = date.fromisoformat(to_date_str)
                    memorization_records = memorization_records.filter(date__range=[from_date, to_date])
                    review_records = review_records.filter(date__range=[from_date, to_date])
                    talqeen_records = talqeen_records.filter(date__range=[from_date, to_date])
                except ValueError:
                    messages.error(request, "Invalid date format.")
                    from_date = None
                    to_date = None
            elif from_date_str:
                try:
                    from_date = date.fromisoformat(from_date_str)
                    memorization_records = memorization_records.filter(date__gte=from_date)
                    review_records = review_records.filter(date__gte=from_date)
                    talqeen_records = talqeen_records.filter(date__gte=from_date)
                except ValueError:
                    messages.error(request, "Invalid date format.")
                    from_date = None
            elif to_date_str:
                try:
                    to_date = date.fromisoformat(to_date_str)
                    memorization_records = memorization_records.filter(date__lte=to_date)
                    review_records = review_records.filter(date__lte=to_date)
                    talqeen_records = talqeen_records.filter(date__lte=to_date)
                except ValueError:
                    messages.error(request, "Invalid date format.")
                    to_date = None

        except Halaqa.DoesNotExist:
            messages.error(request, "Selected Halaqa does not exist.")
            selected_halaqa = None

    context = {
        'halaqat': halaqat,
        'selected_halaqa': selected_halaqa,
        'memorization_records': memorization_records,
        'review_records': review_records,
        'talqeen_records': talqeen_records,
        'from_date': from_date,
        'to_date': to_date,
    }

    return render(request, 'admin/halaqa_report.html', context)

from datetime import date
from datetime import date
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .models import Students, DailyMemorizationTask, DailyReviewTask, DailyTalqeenTask

@login_required
def student_report_single_view(request):
    # التحقق من أن المستخدم لديه صلاحيات الإدارة
    if not request.user.groups.filter(name='admin').exists():
        messages.error(request, "You do not have permission to access this page.")
        return redirect('index')  # أو صفحة مناسبة لعدم الصلاحية

    students = Students.objects.all()
    report_data = None

    if request.method == 'GET':
        student_id = request.GET.get('student')
        from_date_str = request.GET.get('from_date')
        to_date_str = request.GET.get('to_date')

        student = None
        if student_id:
            try:
                student = Students.objects.get(id=student_id)
            except Students.DoesNotExist:
                messages.error(request, "Selected student does not exist.")

        if student:
            memorization_records = DailyMemorizationTask.objects.filter(student=student)
            review_records = DailyReviewTask.objects.filter(student=student)
            talqeen_records = DailyTalqeenTask.objects.filter(student=student)

            # تصفية البيانات حسب التواريخ
            if from_date_str and to_date_str:
                try:
                    from_date = date.fromisoformat(from_date_str)
                    to_date = date.fromisoformat(to_date_str)
                    memorization_records = memorization_records.filter(date__range=[from_date, to_date])
                    review_records = review_records.filter(date__range=[from_date, to_date])
                    talqeen_records = talqeen_records.filter(date__range=[from_date, to_date])
                except ValueError:
                    messages.error(request, "Invalid date format.")
            elif from_date_str:
                try:
                    from_date = date.fromisoformat(from_date_str)
                    memorization_records = memorization_records.filter(date__gte=from_date)
                    review_records = review_records.filter(date__gte=from_date)
                    talqeen_records = talqeen_records.filter(date__gte=from_date)
                except ValueError:
                    messages.error(request, "Invalid date format.")
            elif to_date_str:
                try:
                    to_date = date.fromisoformat(to_date_str)
                    memorization_records = memorization_records.filter(date__lte=to_date)
                    review_records = review_records.filter(date__lte=to_date)
                    talqeen_records = talqeen_records.filter(date__lte=to_date)
                except ValueError:
                    messages.error(request, "Invalid date format.")

            report_data = {
                'memorization': memorization_records,
                'review': review_records,
                'dictation': talqeen_records,
            }

    context = {
        'students': students,
        'report_data': report_data,
    }

    return render(request, 'admin/student_report_single.html', context)


from django.http import HttpResponse
import pandas as pd
import pandas as pd
from django.http import HttpResponse
from datetime import date
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404
from .models import Students, DailyMemorizationTask, DailyReviewTask, DailyTalqeenTask

from django.http import HttpResponse
import pandas as pd

@login_required
def export_student_report_single_excel_view(request, student_id, from_date, to_date):
    print(f"Export view called for student_id: {student_id}, from_date: {from_date}, to_date: {to_date}")
    # Check if the user is an admin
    if not request.user.groups.filter(name='admin').exists():
        messages.error(request, "You do not have permission to access this page.")
        print("User is not an admin.")
        return redirect('index') # Or a suitable unauthorized page

    student = get_object_or_404(Students, id=student_id)
    print(f"Found student: {student.first_name} {student.last_name}")

    from_date_obj = None
    to_date_obj = None

    if from_date and from_date != 'None':
        try:
            from_date_obj = date.fromisoformat(from_date)
            print(f"Parsed from_date: {from_date_obj}")
        except ValueError:
            messages.error(request, "Invalid 'from' date format.")
            print(f"Invalid 'from' date format: {from_date}")
            return redirect('student_report_single') # Redirect back to the report page

    if to_date and to_date != 'None':
        try:
            to_date_obj = date.fromisoformat(to_date)
            print(f"Parsed to_date: {to_date_obj}")
        except ValueError:
            messages.error(request, "Invalid 'to' date format.")
            print(f"Invalid 'to' date format: {to_date}")
            return redirect('student_report_single') # Redirect back to the report page

    memorization_records = DailyMemorizationTask.objects.filter(student=student)
    review_records = DailyReviewTask.objects.filter(student=student)
    talqeen_records = DailyTalqeenTask.objects.filter(student=student)

    # Apply date filtering
    if from_date_obj and to_date_obj:
        memorization_records = memorization_records.filter(date__range=[from_date_obj, to_date_obj])
        review_records = review_records.filter(date__range=[from_date_obj, to_date_obj])
        talqeen_records = talqeen_records.filter(date__range=[from_date_obj, to_date_obj])
        print(f"Filtering by date range: {from_date_obj} to {to_date_obj}")
    elif from_date_obj:
        memorization_records = memorization_records.filter(date__gte=from_date_obj)
        review_records = review_records.filter(date__gte=from_date_obj)
        talqeen_records = talqeen_records.filter(date__gte=from_date_obj)
        print(f"Filtering by from_date: {from_date_obj}")
    elif to_date_obj:
        memorization_records = memorization_records.filter(date__lte=to_date_obj)
        review_records = review_records.filter(date__lte=to_date_obj)
        talqeen_records = talqeen_records.filter(date__lte=to_date_obj)
        print(f"Filtering by to_date: {to_date_obj}")
    else:
        print("No date filtering applied.")


    print(f"Memorization records count: {memorization_records.count()}")
    print(f"Review records count: {review_records.count()}")
    print(f"Talqeen records count: {talqeen_records.count()}")


    # Create a new Excel workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    print("Workbook created.")

    # Memorization Sheet
    mem_sheet = workbook.active
    mem_sheet.title = "تقرير الحفظ"
    mem_sheet.append(["التاريخ", "الفئة", "من سورة", "آية", "إلى سورة", "آية", "عدد الآيات", "عدد الصفحات", "أجزاء الصفحة", "الدرجة"])
    for record in memorization_records:
        mem_sheet.append([
            record.date,
            record.get_MemorizationTask_display(),
            record.previous_surah,
            record.previous_ayat,
            record.current_surah,
            record.current_ayat,
            record.memorized_ayat_count,
            record.pages,
            record.partial_page,
            record.grade,
        ])
    print("Memorization sheet populated.")

    # Review Sheet
    review_sheet = workbook.create_sheet("تقرير المراجعة")
    review_sheet.append(["التاريخ", "الفئة", "من سورة", "آية", "إلى سورة", "آية", "عدد الآيات", "عدد الصفحات", "أجزاء الصفحة", "الدرجة"])
    for record in review_records:
        review_sheet.append([
            record.date,
            record.get_ReviewTask_display(),
            record.previous_surah,
            record.previous_ayat,
            record.current_surah,
            record.current_ayat,
            record.reviewed_ayat_count,
            record.pages,
            record.partial_page,
            record.grade,
        ])
    print("Review sheet populated.")


    # Talqeen Sheet
    talqeen_sheet = workbook.create_sheet("تقرير التلقين")
    talqeen_sheet.append(["التاريخ", "الفئة", "من سورة", "آية", "إلى سورة", "آية", "عدد الآيات", "عدد الصفحات", "أجزاء الصفحة", "الدرجة"])
    for record in talqeen_records:
        talqeen_sheet.append([
            record.date,
            record.get_TalqeenTask_display(),
            record.previous_surah,
            record.previous_ayat,
            record.current_surah,
            record.current_ayat,
            record.memorized_ayat_count, # Assuming memorized_ayat_count is used for Talqeen as well
            record.pages,
            record.partial_page,
            record.grade,
        ])
    print("Talqeen sheet populated.")


    # Prepare the HTTP response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=student_report_{student.first_name}_{student.last_name}.xlsx'
    print("HTTP response prepared.")

    # Save the workbook to the response
    workbook.save(response)
    print("Workbook saved to response.")

    return response




@login_required
def student_report_view(request):
    # Check if the user is an admin
    if not request.user.groups.filter(name='admin').exists():
        messages.error(request, "You do not have permission to access this page.")
        return redirect('index') # Or a suitable unauthorized page

    # Check if the user is an admin
    if not request.user.groups.filter(name='admin').exists():
        messages.error(request, "You do not have permission to access this page.")
        return redirect('index') # Or a suitable unauthorized page

    search_query = request.GET.get('search')

    if search_query:
        students = Students.objects.select_related('halaqa__teacher', 'halaqa__msar', 'halaqa__department').filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(id__icontains=search_query) |
            Q(halaqa__name__icontains=search_query) |
            Q(halaqa__type_class__name__icontains=search_query) |
            Q(halaqa__msar__name__icontains=search_query) |
            Q(halaqa__department__name__icontains=search_query)
        )
        # Prepare data for JSON response
        student_data = []
        for student in students:
            student_data.append({
                'id': student.id,
                'first_name': student.first_name,
                'last_name': student.last_name,
                'halaqa_name': student.halaqa.name if student.halaqa else '',
                'typeread_name': student.halaqa.type_class.name if student.halaqa and student.halaqa.type_class else '',
                'msar_name': student.halaqa.msar.name if student.halaqa and student.halaqa.msar else '',
                'department_name': student.halaqa.department.name if student.halaqa and student.halaqa.department else '',
                'whatsapp': student.whatsapp, # Include whatsapp number
            })
        return JsonResponse({'students': student_data})
    else:
        # Fetch all students with related information for initial load
        students = Students.objects.select_related('halaqa__teacher', 'halaqa__msar', 'halaqa__department').all()
        context = {
            'students': students
        }
        return render(request, 'admin/student_report.html', context)

@login_required
def export_student_report_excel(request):
    # Check if the user is an admin
    if not request.user.groups.filter(name='admin').exists():
        messages.error(request, "You do not have permission to access this page.")
        return redirect('index') # Or a suitable unauthorized page

    # Fetch student data (can apply filtering based on request.GET if needed)
    students = Students.objects.select_related('halaqa__teacher', 'halaqa__msar', 'halaqa__department').all()

    # Create a new Excel workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "تقرير الطلاب"

    # Write headers
    headers = ["اسم الطالب", "رقم الطالب", "اسم الحلقة", "اسم القراءة", "اسم المسار", "اسم القسم", "رقم الواتس"]
    worksheet.append(headers)

    # Write data rows
    for student in students:
        worksheet.append([
            f"{student.first_name} {student.last_name}",
            student.id,
            student.halaqa.name if student.halaqa else '',
            student.halaqa.type_class.name if student.halaqa and student.halaqa.type_class else '',
            student.halaqa.msar.name if student.halaqa and student.halaqa.msar else '',
            student.halaqa.department.name if student.halaqa and student.halaqa.department else '',
            str(student.whatsapp) if student.whatsapp else '', # Include whatsapp number and convert to string
        ])

    # Prepare the HTTP response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=student_report.xlsx'

    # Save the workbook to the response
    workbook.save(response)

    return response

@login_required
def export_halaqa_report_excel(request, halaqa_id):
    # Check if the user is an admin
    if not request.user.groups.filter(name='admin').exists():
        messages.error(request, "You do not have permission to access this page.")
        return redirect('index') # Or a suitable unauthorized page

    halaqa = get_object_or_404(Halaqa, id=halaqa_id)
    students_in_halaqa = Students.objects.filter(halaqa=halaqa)

    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')

    memorization_records = DailyMemorizationTask.objects.filter(student__in=students_in_halaqa)
    review_records = DailyReviewTask.objects.filter(student__in=students_in_halaqa)
    talqeen_records = DailyTalqeenTask.objects.filter(student__in=students_in_halaqa)

    if from_date_str and to_date_str:
        try:
            from_date = date.fromisoformat(from_date_str)
            to_date = date.fromisoformat(to_date_str)
            memorization_records = memorization_records.filter(date__range=[from_date, to_date])
            review_records = review_records.filter(date__range=[from_date, to_date])
            talqeen_records = talqeen_records.filter(date__range=[from_date, to_date])
        except ValueError:
            messages.error(request, "Invalid date format.")
            from_date = None
            to_date = None
    elif from_date_str:
        try:
            from_date = date.fromisoformat(from_date_str)
            memorization_records = memorization_records.filter(date__gte=from_date)
            review_records = review_records.filter(date__gte=from_date)
            talqeen_records = talqeen_records.filter(date__gte=from_date)
        except ValueError:
            messages.error(request, "Invalid date format.")
            from_date = None
    elif to_date_str:
        try:
            to_date = date.fromisoformat(to_date_str)
            memorization_records = memorization_records.filter(date__lte=to_date)
            review_records = review_records.filter(date__lte=to_date)
            talqeen_records = talqeen_records.filter(date__lte=to_date)
        except ValueError:
            messages.error(request, "Invalid date format.")
            to_date = None

    # Create a new Excel workbook and select the active worksheet
    workbook = openpyxl.Workbook()

    # Memorization Sheet
    mem_sheet = workbook.active
    mem_sheet.title = "تقرير الحفظ"
    mem_sheet.append(["الطالب", "التاريخ", "الفئة", "من سورة", "آية", "إلى سورة", "آية", "عدد الصفحات", "جزء من صفحة", "الدرجة", "عدد الآيات المحفوظة"])
    for record in memorization_records:
        mem_sheet.append([
            f"{record.student.first_name} {record.student.last_name}",
            record.date,
            record.get_MemorizationTask_display(),
            record.previous_surah,
            record.previous_ayat,
            record.current_surah,
            record.current_ayat,
            record.pages,
            record.partial_page,
            record.grade,
            record.memorized_ayat_count,
        ])

    # Review Sheet
    review_sheet = workbook.create_sheet("تقرير المراجعة")
    review_sheet.append(["الطالب", "التاريخ", "الفئة", "من سورة", "آية", "إلى سورة", "آية", "عدد الصفحات", "جزء من صفحة", "الدرجة", "عدد الآيات المراجعة"])
    for record in review_records:
        review_sheet.append([
            f"{record.student.first_name} {record.student.last_name}",
            record.date,
            record.get_ReviewTask_display(),
            record.previous_surah,
            record.previous_ayat,
            record.current_surah,
            record.current_ayat,
            record.pages,
            record.partial_page,
            record.grade,
            record.reviewed_ayat_count,
        ])

    # Talqeen Sheet
    talqeen_sheet = workbook.create_sheet("تقرير التلقين")
    talqeen_sheet.append(["الطالب", "التاريخ", "الفئة", "من سورة", "آية", "إلى سورة", "آية", "عدد الصفحات", "جزء من صفحة", "الدرجة", "عدد الآيات الملقنة"])
    for record in talqeen_records:
        talqeen_sheet.append([
            f"{record.student.first_name} {record.student.last_name}",
            record.date,
            record.get_TalqeenTask_display(),
            record.previous_surah,
            record.previous_ayat,
            record.current_surah,
            record.current_ayat,
            record.pages,
            record.partial_page,
            record.grade,
            record.memorized_ayat_count,
        ])


    # Prepare the HTTP response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=halaqa_report_{halaqa.name}.xlsx'

    # Save the workbook to the response
    workbook.save(response)

    return response


@login_required
def teacher_report_view(request):
    # Check if the user is an admin
    if not request.user.groups.filter(name='admin').exists():
        messages.error(request, "You do not have permission to access this page.")
        return redirect('index') # Or a suitable unauthorized page

    search_query = request.GET.get('search')

    if search_query:
        # Filter Teachers objects based on their fields or related Halaqa fields
        teachers = Teachers.objects.select_related('user').filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(gender__icontains=search_query) |
            Q(nationality__icontains=search_query) |
            Q(mobile_whatsapp__icontains=search_query) |
            Q(halaqa__name__icontains=search_query) |
            Q(halaqa__type_class__name__icontains=search_query) |
            Q(halaqa__msar__name__icontains=search_query) |
            Q(halaqa__department__name__icontains=search_query)
        ).distinct() # Use distinct to avoid duplicate teachers if they are in multiple halaqas (though typically a teacher is in one)

        # Prepare data for JSON response
        teacher_data = []
        for teacher in teachers:
            # Get the halaqa associated with the teacher, if any
            halaqa = Halaqa.objects.filter(teacher=teacher).first()
            teacher_data.append({
                'id': teacher.id,
                'first_name': teacher.first_name,
                'last_name': teacher.last_name,
                'gender': teacher.gender,
                'nationality': teacher.nationality,
                'whatsapp': teacher.mobile_whatsapp,
                'halaqa_name': halaqa.name if halaqa else 'غير مرتبط بحلقة', # Display 'غير مرتبط بحلقة' if no halaqa
                'typeread_name': halaqa.type_class.name if halaqa and halaqa.type_class else '',
                'msar_name': halaqa.msar.name if halaqa and halaqa.msar else '',
                'department_name': halaqa.department.name if halaqa and halaqa.department else '',
            })
        return JsonResponse({'teachers': teacher_data})
    else:
        # Fetch all Teachers objects for initial load
        teachers = Teachers.objects.select_related('user').all()

        teacher_data = []
        for teacher in teachers:
            # Get the halaqa associated with the teacher, if any
            halaqa = Halaqa.objects.filter(teacher=teacher).first()
            teacher_data.append({
                'id': teacher.id,
                'first_name': teacher.first_name,
                'last_name': teacher.last_name,
                'gender': teacher.gender,
                'nationality': teacher.nationality,
                'whatsapp': teacher.mobile_whatsapp,
                'halaqa_name': halaqa.name if halaqa else 'غير مرتبط بحلقة', # Display 'غير مرتبط بحلقة' if no halaqa
                'typeread_name': halaqa.type_class.name if halaqa and halaqa.type_class else '',
                'msar_name': halaqa.msar.name if halaqa and halaqa.msar else '',
                'department_name': halaqa.department.name if halaqa and halaqa.department else '',
            })

        context = {
            'teachers': teacher_data # Pass the list of teacher data
        }
        return render(request, 'admin/teacher_report.html', context)

@login_required
def export_teacher_report_excel(request):
    # Check if the user is an admin
    if not request.user.groups.filter(name='admin').exists():
        messages.error(request, "You do not have permission to access this page.")
        return redirect('index') # Or a suitable unauthorized page

    # Fetch all Teachers objects
    teachers = Teachers.objects.select_related('user').all()

    # Create a new Excel workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "تقرير المدرسين"

    # Write headers
    headers = ["اسم المدرس", "الجنس", "الجنسية", "رقم الواتس الخاص بالمدرس", "اسم الحلقة", "اسم القراءة", "اسم المسار", "اسم القسم"]
    worksheet.append(headers)

    # Write data rows
    for teacher in teachers:
        # Get the halaqa associated with the teacher, if any
        halaqa = Halaqa.objects.filter(teacher=teacher).first()
        worksheet.append([
            f"{teacher.first_name} {teacher.last_name}",
            teacher.gender,
            teacher.nationality,
            str(teacher.mobile_whatsapp) if teacher.mobile_whatsapp else '', # Convert PhoneNumber to string
            halaqa.name if halaqa else 'غير مرتبط بحلقة', # Display 'غير مرتبط بحلقة' if no halaqa
            halaqa.type_class.name if halaqa and halaqa.type_class else '',
            halaqa.msar.name if halaqa and halaqa.msar else '',
            halaqa.department.name if halaqa and halaqa.department else '',
        ])

    # Prepare the HTTP response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=teacher_report.xlsx'

    # Save the workbook to the response
    workbook.save(response)

    return response

@login_required
@csrf_exempt # Add csrf_exempt for simplicity during development, consider proper CSRF handling in production
def export_students_excel_view(request):
    # Check if the user is an admin
    if not request.user.groups.filter(name='admin').exists():
        messages.error(request, "You do not have permission to access this page.")
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            students_data = data.get('students', [])

            if not students_data:
                return JsonResponse({'error': 'No student data provided'}, status=400)

            # Create a new Excel workbook and select the active worksheet
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "تقرير الطلاب"

            # Write headers (assuming all student objects have the same keys)
            if students_data:
                headers = list(students_data[0].keys())
                worksheet.append(headers)

                # Write data rows
                for student in students_data:
                    row_data = [student.get(header, '') for header in headers]
                    worksheet.append(row_data)

            # Prepare the HTTP response
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=students_report.xlsx'

            # Save the workbook to the response
            workbook.save(response)

            return response

        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'Invalid request method'}, status=405)

@login_required
def memorization_report_view(request):
    # Check if the user is an admin
    if not request.user.groups.filter(name='admin').exists():
        messages.error(request, "You do not have permission to access this page.")
        return redirect('index') # Or a suitable unauthorized page

    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')

    memorization_tasks = DailyMemorizationTask.objects.all()

    from_date = None
    to_date = None

    if from_date_str and to_date_str:
        try:
            from_date = date.fromisoformat(from_date_str)
            to_date = date.fromisoformat(to_date_str)
            memorization_tasks = memorization_tasks.filter(date__range=[from_date, to_date])
        except ValueError:
            messages.error(request, "Invalid date format.")
            # If date format is invalid, proceed with all data or handle as needed
            pass
    elif from_date_str:
         try:
            from_date = date.fromisoformat(from_date_str)
            memorization_tasks = memorization_tasks.filter(date__gte=from_date)
         except ValueError:
            messages.error(request, "Invalid date format.")
            pass
    elif to_date_str:
        try:
            to_date = date.fromisoformat(to_date_str)
            memorization_tasks = memorization_tasks.filter(date__lte=to_date)
        except ValueError:
            messages.error(request, "Invalid date format.")
            pass

# Calculate overall total memorization for the main table
    overall_total_pages = DailyMemorizationTask.objects.aggregate(Sum('pages'))['pages__sum'] or 0
    overall_total_partial_pages = DailyMemorizationTask.objects.aggregate(Sum('partial_page'))['partial_page__sum'] or 0
    total_pages_and_parts = overall_total_pages + overall_total_partial_pages

    # Calculate total memorization
    total_memorization = memorization_tasks.aggregate(
        total_ayat=Coalesce(Sum('memorized_ayat_count', output_field=IntegerField()), 0),
        total_pages=Coalesce(Sum('pages', output_field=IntegerField()), 0),
        total_partial_pages=Coalesce(Sum('partial_page', output_field=IntegerField()), 0)
    )

    # Calculate memorization by category
    memorization_by_category = memorization_tasks.values('MemorizationTask').annotate(
        total_ayat=Coalesce(Sum('memorized_ayat_count', output_field=IntegerField()), 0),
        total_pages=Coalesce(Sum('pages', output_field=IntegerField()), 0),
        total_partial_pages=Coalesce(Sum('partial_page'), 0.0), # Remove IntegerField cast and default to float
        total_pages_and_parts=Coalesce(Sum('pages', output_field=IntegerField()), 0) + Coalesce(Sum('partial_page'), 0.0) # Calculate total pages and parts with float sum
    ).order_by('MemorizationTask')

    # Get the category choices from the DailyMemorizationTask model
    category_choices = dict(DailyMemorizationTask.حفظ_الفئة)

    # Add category name to each item in memorization_by_category
    for category_data in memorization_by_category:
        category_number = category_data['MemorizationTask']
        category_data['category_name'] = category_choices.get(category_number, 'Unknown Category')


    # Calculate memorization by halaqa
    memorization_by_halaqa = memorization_tasks.values(
        'student__halaqa__department__name',
        'student__halaqa__msar__name',
        'student__halaqa__type_class__name',
        'student__halaqa__name'
    ).annotate(
        total_ayat=Coalesce(Sum('memorized_ayat_count', output_field=IntegerField()), 0),
        total_pages=Coalesce(Sum('pages', output_field=IntegerField()), 0),
        total_partial_pages=Coalesce(Sum('partial_page'), 0.0), # Remove IntegerField cast and default to float
        total_pages_and_parts=Coalesce(Sum('pages', output_field=IntegerField()), 0) + Coalesce(Sum('partial_page'), 0.0) # Calculate total pages and parts
    ).order_by(
        'student__halaqa__department__name',
        'student__halaqa__msar__name',
        'student__halaqa__type_class__name',
        'student__halaqa__name'
    )

    # Calculate memorization by student
    memorization_by_student = memorization_tasks.values('student__id', 'student__first_name', 'student__last_name').annotate(
        total_ayat=Coalesce(Sum('memorized_ayat_count', output_field=IntegerField()), 0),
        total_pages=Coalesce(Sum('pages', output_field=IntegerField()), 0),
        total_partial_pages=Coalesce(Sum('partial_page'), 0.0), # Remove IntegerField cast and default to float
        total_pages_and_parts=Coalesce(Sum('pages', output_field=IntegerField()), 0) + Coalesce(Sum('partial_page'), 0.0) # Calculate total pages and parts
    ).order_by('student__first_name', 'student__last_name')

    # Combine first and last name for student report
    student_report_data = []
    for student_data in memorization_by_student:
        student_report_data.append({
            'student_id': student_data['student__id'],
            'student_name': f"{student_data['student__first_name']} {student_data['student__last_name']}",
            'total_ayat': student_data['total_ayat'],
            'total_pages': student_data['total_pages'],
            'total_partial_pages': student_data['total_partial_pages'],
            'total_pages_and_parts': student_data['total_pages_and_parts'], # Add total pages and parts
        })


    context = {
        'from_date': from_date,
        'to_date': to_date,
        'total_memorized_ayat': total_memorization['total_ayat'],
        'total_pages': total_memorization['total_pages'],
        'total_partial_pages': total_memorization['total_partial_pages'],
        'total_pages_and_parts': total_pages_and_parts,
        'memorization_by_category': memorization_by_category,
        'memorization_by_halaqa': memorization_by_halaqa,
        'memorization_by_student': student_report_data,
    }

    # Get the category choices from the DailyMemorizationTask model
    category_choices = dict(DailyMemorizationTask.حفظ_الفئة)

    context = {
        'from_date': from_date,
        'to_date': to_date,
        'total_memorized_ayat': total_memorization['total_ayat'],
        'total_pages': total_memorization['total_pages'],
        'total_partial_pages': total_memorization['total_partial_pages'],
        'total_pages_and_parts': total_pages_and_parts,
        'memorization_by_category': memorization_by_category,
        'memorization_by_halaqa': memorization_by_halaqa,
        'memorization_by_student': student_report_data,
        'category_choices': category_choices, # Add category choices to context
    }

    return render(request, 'admin/memorization_report.html', context)


@login_required
def manage_sections(request):
    # Check if the user is an admin
    if not request.user.groups.filter(name='admin').exists():
        messages.error(request, "You do not have permission to access this page.")
        return redirect('index') # Or a suitable unauthorized page

    # Fetch all departments from the database
    departments = Department.objects.all()

    # Pass the departments to the template context
    context = {
        'departments': departments
    }

    return render(request, 'manage/sections_manager.html', context)


@login_required
def manage_msars(request):
    # Check if the user is an admin
    if not request.user.groups.filter(name='admin').exists():
        messages.error(request, "You do not have permission to access this page.")
        return redirect('index') # Or a suitable unauthorized page

    # Fetch all paths from the database
    paths = Msar.objects.all()

    # Pass the paths to the template context
    context = {
        'paths': paths
    }

    return render(request, 'manage/manage_msars.html', context)
@login_required
def manage_halaqat(request):
    # Check if the user is an admin
    if not request.user.groups.filter(name='admin').exists():
        messages.error(request, "You do not have permission to access this page.")
        return redirect('index') # Or a suitable unauthorized page

    # Fetch all halaqat from the database
    halaqat = Halaqa.objects.all()

    # Pass the halaqat to the template context
    context = {
        'halaqat': halaqat
    }

    return render(request, 'manage/manage_halaqa.html', context)


@login_required
def student_distribution_view(request):
    # Check if the user is an admin
    if not request.user.groups.filter(name='admin').exists():
        messages.error(request, "You do not have permission to access this page.")
        return redirect('index') # Or a suitable unauthorized page

    # Fetch necessary data for the template (e.g., students, halaqat)
    students = Students.objects.all()
    halaqat = Halaqa.objects.all()
    total_students = Students.objects.count() # Calculate total number of students

    # Serialize halaqat data to JSON
    halaqat_list = list(halaqat.values('id', 'name'))
    import json
    halaqat_json = json.dumps(halaqat_list)

    context = {
        'students': students,
        'halaqat': halaqat, # Keep the original for the select options
        'halaqat_json': halaqat_json, # Add JSON data for JavaScript
        'total_students': total_students, # Add total student count to context
    }

    return render(request, 'manage/student_distribution.html', context)


@login_required
def manage_typereads(request):
    # Check if the user is an admin
    if not request.user.groups.filter(name='admin').exists():
        messages.error(request, "You do not have permission to access this page.")
        return redirect('index') # Or a suitable unauthorized page

    # Fetch all typereads from the database
    typereads = TypeRead.objects.all()

    # Pass the typereads to the template context
    context = {
        'typereads': typereads
    }

    return render(request, 'manage/manage_typeread.html', context)

@login_required
def manage_students(request):
    # Check if the user is an admin
    if not request.user.groups.filter(name='admin').exists():
        messages.error(request, "You do not have permission to access this page.")
        return redirect('index') # Or a suitable unauthorized page

    search_query = request.GET.get('q')
    students = Students.objects.all()

    if search_query:
        students = students.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query)
        )

    # Pass the students to the template context
    context = {
        'students': students
    }

    return render(request, 'admin/manage_students.html', context)

@login_required
@csrf_exempt # Consider more secure CSRF handling in production
def import_students_view(request):
    # Check if the user is an admin
    if not request.user.groups.filter(name='admin').exists():
        messages.error(request, "You do not have permission to access this page.")
        return redirect('index') # Or a suitable unauthorized page

    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = form.cleaned_data['excel_file']

            try:
                # Load the workbook
                workbook = openpyxl.load_workbook(excel_file)
                sheet = workbook.active # Get the active worksheet

                # Assume the first row is the header
                header = [cell.value for cell in sheet[1]]
                # Map header names to model fields (adjust these mappings based on your Excel file)
                # Example: {'الاسم الأول': 'first_name', 'اسم المستخدم': 'username', ...}
                # You will need to define the exact mapping based on the expected Excel columns.
                # For now, I'll use placeholder names. You MUST adjust this mapping.
                column_mapping = {
                    'الاسم الأول': 'first_name',
                    'الاسم الثاني': 'second_name',
                    'الاسم الأخير': 'last_name',
                    'تاريخ الميلاد': 'birthday',
                    'الجنسية': 'nationality',
                    'الجنس': 'gender',
                    'البريد الإلكتروني': 'email',
                    'اللغة': 'language', # Assuming language name in Excel
                    'رقم الجوال': 'mobile',
                    'رقم الواتساب': 'whatsapp',
                    'اسم المستخدم': 'username',
                    'كلمة المرور': 'password', # Assuming plain text password in Excel (NOT recommended)
                    'الحلقة': 'halaqa', # Assuming halaqa name in Excel
                    'تفعيل': 'activate', # Assuming boolean value (True/False, 1/0, Yes/No)
                }

                # Validate that required columns exist in the header
                required_columns = ['الاسم الأول', 'اسم المستخدم', 'كلمة المرور'] # Add other required columns
                if not all(col in header for col in required_columns):
                     messages.error(request, "ملف الاكسل لا يحتوي على جميع الأعمدة المطلوبة.")
                     return redirect('manage_students')


                # Start an atomic transaction
                imported_count = 0
                errors = []
                warnings = []

                with transaction.atomic():
                    for row_index in range(2, sheet.max_row + 1): # Start from the second row (after header)
                        row_data = {}
                        try:
                            for col_index, cell in enumerate(sheet[row_index]):
                                # Use the header to get the column name and map it to the model field
                                column_name = header[col_index]
                                if column_name in column_mapping:
                                    row_data[column_mapping[column_name]] = cell.value

                            # Extract data using the mapped keys
                            username = row_data.get('username')
                            password = row_data.get('password') # Again, storing plain text is NOT recommended
                            first_name = row_data.get('first_name', '')
                            last_name = row_data.get('last_name', '')
                            email = row_data.get('email')
                            birthday = row_data.get('birthday')
                            nationality = row_data.get('nationality', '')
                            gender = row_data.get('gender', '')
                            language_name = row_data.get('language')
                            mobile = row_data.get('mobile')
                            whatsapp = row_data.get('whatsapp')
                            halaqa_name = row_data.get('halaqa')
                            # Handle 'تفعيل' column: convert various inputs to boolean
                            activate_value = str(row_data.get('activate', '')).lower()
                            if activate_value in ['true', 'صحيح', '1']:
                                activate = True
                            elif activate_value in ['false', 'خطأ', '0']:
                                activate = False
                            else:
                                # Default to True if value is not recognized
                                activate = True
                                warnings.append(f"الصف {row_index}: تم استخدام القيمة الافتراضية (تفعيل) في عمود 'تفعيل' بسبب قيمة غير معروفة: {row_data.get('activate')}")

                            if not username or not password:
                                warnings.append(f"الصف {row_index}: تم تخطي الصف بسبب اسم المستخدم أو كلمة المرور مفقودة.")
                                continue # Skip this row if essential data is missing

                            # Find or create the Django User
                            user, created = User.objects.get_or_create(username=username)
                            if created:
                                user.set_password(password) # Hash the password
                                user.first_name = first_name
                                user.last_name = last_name
                                user.email = email
                                user.save()
                                warnings.append(f"الصف {row_index}: تم إنشاء مستخدم جديد لـ {username}.")
                            else:
                                # If user exists, update their details (optional, based on requirements)
                                user.first_name = first_name
                                user.last_name = last_name
                                if email: # Only update email if provided in the Excel
                                    user.email = email
                                # Note: Updating password for existing users from Excel is highly discouraged
                                user.save()
                                warnings.append(f"الصف {row_index}: تم تحديث بيانات المستخدم الحالي لـ {username}.")


                            # Ensure the user is in the 'students' group
                            try:
                                students_group, created_group = Group.objects.get_or_create(name='students')
                                if created_group:
                                    warnings.append(f"تم إنشاء مجموعة 'students' لأنها لم تكن موجودة.")
                                user.groups.add(students_group)
                                warnings.append(f"الصف {row_index}: تم إضافة المستخدم {username} إلى مجموعة 'students'.")
                            except Exception as e:
                                errors.append(f"الصف {row_index}: حدث خطأ أثناء إضافة المستخدم {username} إلى مجموعة 'students': {e}")
                                continue # Skip this row on error

                            # Find or create the Student profile
                            student, created = Students.objects.get_or_create(user=user, defaults={'username': username})

                            # Update student details
                            student.first_name = first_name
                            student.second_name = row_data.get('second_name', '') # Ensure second_name is handled
                            student.last_name = last_name

                            # Handle Birthday: Attempt to parse various formats
                            birthday_value = row_data.get('birthday')
                            parsed_birthday = None
                            if birthday_value:
                                if isinstance(birthday_value, date):
                                    # If openpyxl already read it as a date
                                    parsed_birthday = birthday_value
                                elif isinstance(birthday_value, str):
                                    # Attempt to parse string formats (e.g., 'YYYY-MM-DD', 'MM/DD/YYYY', 'DD/MM/YYYY')
                                    try:
                                        # Try YYYY-MM-DD
                                        parsed_birthday = date.fromisoformat(birthday_value)
                                    except ValueError:
                                        try:
                                            # Try MM/DD/YYYY
                                            parsed_birthday = datetime.strptime(birthday_value, '%m/%d/%Y').date()
                                        except ValueError:
                                            try:
                                                # Try DD/MM/YYYY
                                                parsed_birthday = datetime.strptime(birthday_value, '%d/%m/%Y').date()
                                            except ValueError:
                                                # If none of the common formats match
                                                errors.append(f"الصف {row_index}: تنسيق تاريخ الميلاد غير صالح: {birthday_value}")
                                                continue # Skip this row if birthday format is invalid
                                # Add more elif blocks here for other potential string formats if needed
                                else:
                                    # Handle other types if necessary (e.g., numbers representing dates)
                                    errors.append(f"الصف {row_index}: نوع بيانات تاريخ الميلاد غير مدعوم: {type(birthday_value)}")
                                    continue # Skip this row if birthday data type is unsupported

                            if parsed_birthday is None:
                                errors.append(f"الصف {row_index}: تاريخ الميلاد مفقود أو غير صالح.")
                                continue # Skip this row if birthday is missing or invalid

                            student.birthday = parsed_birthday
                            student.nationality = nationality
                            student.gender = gender
                            student.email = email
                            student.mobile = mobile
                            student.whatsapp = whatsapp
                            student.activate = activate

                            # Handle Language (assuming language_name is the language name string)
                            if language_name:
                                try:
                                    language_obj = Language.objects.get(name=language_name)
                                    student.language = language_obj
                                except Language.DoesNotExist:
                                    warnings.append(f"الصف {row_index}: اللغة '{language_name}' للطالب {username} غير موجودة. تم تخطي تعيين اللغة.")
                                    student.language = None # Set to None if language not found
                            else:
                                student.language = None # Ensure language is set to None if not provided


                            # Handle Halaqa (assuming halaqa_name is the halaqa name string)
                            if halaqa_name:
                                try:
                                    halaqa_obj = Halaqa.objects.get(name=halaqa_name)
                                    student.halaqa = halaqa_obj
                                except Halaqa.DoesNotExist:
                                    warnings.append(f"الصف {row_index}: الحلقة '{halaqa_name}' للطالب {username} غير موجودة. تم تخطي تعيين الحلقة.")
                                    student.halaqa = None # Set to None if halaqa not found
                            else:
                                 student.halaqa = None # Ensure halaqa is set to None if not provided

                            # Add an explicit check for birthday before saving
                            if student.birthday is None:
                                errors.append(f"الصف {row_index}: لا يمكن حفظ الطالب '{username}' لأن تاريخ الميلاد مفقود أو غير صالح.")
                                continue # Skip saving this row

                            student.save()

                            # Create or update UserRole for the imported student
                            try:
                                user_role, created_role = UserRole.objects.get_or_create(user=user)
                                user_role.role = 'student'
                                user_role.save()
                                warnings.append(f"الصف {row_index}: تم تعيين دور 'student' للمستخدم {username}.")
                            except Exception as e:
                                errors.append(f"الصف {row_index}: حدث خطأ أثناء تعيين دور 'student' للمستخدم {username}: {e}")
                                # Decide whether to continue or break if role assignment fails
                                continue # Skip this row on error


                            imported_count += 1 # Increment count for successfully processed rows

                        except Exception as e:
                            # Catch any other unexpected errors during row processing
                            errors.append(f"الصف {row_index}: حدث خطأ غير متوقع أثناء معالجة البيانات: {e}")
                            # Continue to the next row

                # Display summary messages after the transaction
                if imported_count > 0:
                    messages.success(request, f"تم استيراد بيانات {imported_count} طالب بنجاح!")
                else:
                    messages.info(request, "لم يتم استيراد أي بيانات طلاب.")

                for warning in warnings:
                    messages.warning(request, warning)

                for error in errors:
                    messages.error(request, error)

                return redirect('manage_students')

            except FileNotFoundError:
                messages.error(request, "حدث خطأ: لم يتم العثور على الملف.")
            except Exception as e:
                messages.error(request, f"حدث خطأ أثناء معالجة الملف: {e}")

        else:
            messages.error(request, "حدث خطأ في النموذج. يرجى التأكد من اختيار ملف.")

    # If not a POST request or form is invalid, redirect back to the manage students page
    return redirect('manage_students')
@login_required
def generate_student_excel_template_view(request):
    # Check if the user is an admin
    if not request.user.groups.filter(name='admin').exists():
        messages.error(request, "You do not have permission to access this page.")
        return redirect('index') # Or a suitable unauthorized page

    # Define the column headers based on the expected import format
    # These should match the keys in the column_mapping in import_students_view
    headers = [
        'الاسم الأول',
        'الاسم الثاني',
        'الاسم الأخير',
        'تاريخ الميلاد',
        'الجنسية',
        'الجنس',
        'البريد الإلكتروني',
        'اللغة',
        'رقم الجوال',
        'رقم الواتساب',
        'اسم المستخدم',
        'كلمة المرور',
        'الحلقة',
        'تفعيل',
    ]

    # Create a new Excel workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "نموذج استيراد الطلاب"

    # Write the headers to the first row
    worksheet.append(headers)

    # Prepare the HTTP response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=student_import_template.xlsx'

    # Save the workbook to the response
    workbook.save(response)

    return response


@login_required
def manage_teachers_view(request):
    # Check if the user is an admin
    if not request.user.groups.filter(name='admin').exists():
        messages.error(request, "You do not have permission to access this page.")
        return redirect('index') # Or a suitable unauthorized page

    # Fetch all teachers
    all_teachers = Teachers.objects.select_related('user').all()

    # Prepare a list of dictionaries with the required data
    teachers_data = []
    for teacher in all_teachers:
        # Get the first halaqa associated with the teacher, if any
        halaqa = teacher.halaqa_set.first()
        msar_name = halaqa.msar.name if halaqa and halaqa.msar else 'غير محدد'

        teachers_data.append({
            'id': teacher.id,
            'first_name': teacher.first_name,
            'last_name': teacher.last_name,
            'mobile_whatsapp': teacher.mobile_whatsapp,
            'username': teacher.user.username,
            'msar_name': msar_name,
        })

    # Pass the prepared data to the template context
    context = {
        'teachers': teachers_data
    }

    return render(request, 'admin/manage_teachers.html', context)
@login_required
def add_teacher_view(request):
    # Check if the user is an admin
    if not request.user.groups.filter(name='admin').exists():
        messages.error(request, "You do not have permission to access this page.")
        return redirect('index') # Or a suitable unauthorized page

    if request.method == 'POST':
        form = TeacherForm(request.POST, request.FILES)
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            email = form.cleaned_data.get('email') or None
            first_name = form.cleaned_data['first_name']
            last_name = form.cleaned_data['last_name']

            # Check if username already exists
            if User.objects.filter(username=username).exists():
                messages.error(request, "اسم المستخدم موجود بالفعل. يرجى اختيار اسم مستخدم آخر.")
                return render(request, 'admin/add_teacher.html', {'form': form})

            # Create a Django User object
            user = User.objects.create_user(
                username=username,
                password=password,
                email=email,
                first_name=first_name,
                last_name=last_name
            )
            user.save()

            # Create the Teacher instance and link it to the created User
            teacher = form.save(commit=False)
            teacher.user = user
            teacher.activate = request.POST.get('activate') == 'on' # Handle activate checkbox
            teacher.save() # Save the teacher instance to create the primary key

            # Handle multiple language selection and set default to Arabic if none selected
            selected_languages = form.cleaned_data.get('language')
            if not selected_languages:
                # If no languages are selected, default to Arabic
                arabic_language, created = Language.objects.get_or_create(name='Arabic')
                teacher.language.set([arabic_language])
            else:
                # Set the selected languages, ensuring it's an iterable
                teacher.language.set(list(selected_languages))


            # Add the user to the "Teachers" group
            try:
                teachers_group = Group.objects.get(name='Teachers')
                user.groups.add(teachers_group)
            except Group.DoesNotExist:
                error_message = "The 'Teachers' group does not exist. Please create it in the Django admin."
                messages.error(request, error_message)
            except Exception as e:
                messages.warning(request, "Could not add user to 'Teachers' group due to an unexpected error.")


            # Send WhatsApp notification (optional, based on requirements)
            whatsapp_number = teacher.mobile_whatsapp
            if whatsapp_number:
                try:
                    # Assuming send_registration_success_whatsapp is appropriate here
                    send_registration_success_whatsapp(str(whatsapp_number))
                except Exception as e:
                    messages.warning(request, "Could not send WhatsApp notification.")


            messages.success(request, "تم إضافة المدرس بنجاح!")
            return render(request, 'admin/add_teacher.html', {'form': form})

        else:
            messages.error(request, "حدث خطأ أثناء إضافة المدرس. يرجى مراجعة الأخطاء أدناه.")
            # The form with errors is rendered here
            return render(request, 'admin/add_teacher.html', {'form': form})
    else:
        form = TeacherForm()

    return render(request, 'admin/add_teacher.html', {'form': form})

@login_required
def edit_teacher_view(request, teacher_id):
    # Check if the user is an admin
    if not request.user.groups.filter(name='admin').exists():
        messages.error(request, "You do not have permission to access this page.")
        return redirect('index') # Or a suitable unauthorized page

    teacher = get_object_or_404(Teachers, id=teacher_id)
    user = teacher.user # Get the associated User object

    if request.method == 'POST':
        # Pass instance to the form to update the existing teacher and user
        form = TeacherForm(request.POST, request.FILES, instance=teacher)
        if form.is_valid():
            # Update the User object fields manually from the form's cleaned data
            user.first_name = form.cleaned_data['first_name']
            user.last_name = form.cleaned_data['last_name']
            user.email = form.cleaned_data.get('email') or '' # Handle optional email, set to empty string if None
            # Note: Username and password changes would require additional logic
            user.save()

            # Save the teacher instance (which is already linked to the user)
            teacher = form.save(commit=False)
            teacher.activate = request.POST.get('activate') == 'on' # Handle activate checkbox
            teacher.save()

            # Handle multiple language selection
            selected_languages = form.cleaned_data.get('language')
            if selected_languages is not None: # Check if language was in the form data
                 teacher.language.set(selected_languages)


            messages.success(request, "تم تحديث بيانات المدرس بنجاح!")
            return redirect('manage_teachers') # Redirect back to the manage teachers page
        else:
            messages.error(request, "حدث خطأ أثناء تحديث بيانات المدرس. يرجى مراجعة الأخطاء أدناه.")
            # Render the form with errors
            return render(request, 'admin/edit_teacher.html', {'form': form, 'teacher': teacher})
    else:
        # Populate the form with the existing teacher and user data
        # Need to pass initial data for User fields to the TeacherForm
        initial_data = {
            'first_name': user.first_name,
            'last_name': user.last_name,
            'email': user.email,
            'username': user.username, # Display username but it won't be editable by default
            # Add other fields from the Teacher model
            'birthday': teacher.birthday,
            'nationality': teacher.nationality,
            'gender': teacher.gender,
            'mobile_whatsapp': teacher.mobile_whatsapp,
            'cv_teacher': teacher.cv_teacher,
            'activate': teacher.activate,
            'language': teacher.language.all(), # Get all languages for initial data
        }
        form = TeacherForm(instance=teacher, initial=initial_data)

    return render(request, 'admin/edit_teacher.html', {'form': form, 'teacher': teacher})


@login_required
def add_student_view(request):
    # Check if the user is an admin
    if not request.user.groups.filter(name='admin').exists():
        messages.error(request, "You do not have permission to access this page.")
        return redirect('index') # Or a suitable unauthorized page

    if request.method == 'POST':
        form = StudentForm(request.POST, request.FILES)
        print("Is form valid?", form.is_valid()) # Debug print
        if form.is_valid():
            user_data = form.cleaned_data
            username = user_data['username']
            password = user_data['password']
            email = user_data.get('email') or None
            first_name = user_data['first_name']
            last_name = user_data['last_name']
            halaqa = user_data.get('halaqa') # Get halaqa from cleaned data

            # Check if username already exists
            if User.objects.filter(username=username).exists():
                messages.error(request, "اسم المستخدم موجود بالفعل. يرجى اختيار اسم مستخدم آخر.")
                return render(request, 'admin/add_student.html', {'form': form})

            # Create a Django User object
            user = User.objects.create_user(
                username=username,
                password=password,
                email=email,
                first_name=first_name,
                last_name=last_name
            )
            user.save()

            # Create the Student instance and link it to the created User
            student = form.save(commit=False)
            student.user = user
            student.activate = request.POST.get('activate') == 'on' # Handle activate checkbox
            student.halaqa = halaqa # Assign the selected halaqa (can be None)
            student.save() # Save the student instance

            # Add the user to the "students" group
            try:
                students_group = Group.objects.get(name='students')
                user.groups.add(students_group)
                print(f"User {user.username} added to 'students' group.") # Debug print
            except Group.DoesNotExist:
                error_message = "The 'students' group does not exist. Please create it in the Django admin."
                print(error_message) # Debug print
                messages.error(request, error_message)
            except Exception as e:
                print(f"Error adding user to 'students' group: {e}") # Debug print
                messages.warning(request, "Could not add user to 'students' group due to an unexpected error.")


            # Send WhatsApp notification (optional, based on requirements)
            whatsapp_number = student.whatsapp
            if whatsapp_number:
                try:
                    # Assuming send_registration_success_whatsapp is appropriate here
                    send_registration_success_whatsapp(str(whatsapp_number))
                    print(f"WhatsApp notification sent to {whatsapp_number}") # Debug print
                except Exception as e:
                    print(f"Error sending WhatsApp message: {e}") # Debug print
                    messages.warning(request, "Could not send WhatsApp notification.")


            messages.success(request, "تم إضافة الطالب بنجاح!")
            return redirect('manage_students') # Redirect to the manage students page after adding

        else:
            print("Form errors:", form.errors) # Debug print
            messages.error(request, "حدث خطأ أثناء إضافة الطالب. يرجى مراجعة الأخطاء أدناه.")
            # The form with errors is rendered here
            return render(request, 'admin/add_student.html', {'form': form})
    else:
        form = StudentForm()

    return render(request, 'admin/add_student.html', {'form': form})


@login_required
def edit_student_view(request, student_id):
    # Check if the user is an admin
    if not request.user.groups.filter(name='admin').exists():
        messages.error(request, "You do not have permission to access this page.")
        return redirect('index') # Or a suitable unauthorized page

    student = get_object_or_404(Students, id=student_id)
    user = student.user # Get the associated User object

    if request.method == 'POST':
        # Pass instance to the form to update the existing student and user
        form = StudentForm(request.POST, request.FILES, instance=student)
        if form.is_valid():
            # Update the User object fields manually from the form's cleaned data
            user.first_name = form.cleaned_data['first_name']
            user.last_name = form.cleaned_data['last_name']
            user.email = form.cleaned_data.get('email') or '' # Handle optional email, set to empty string if None
            # Note: Username and password changes would require additional logic
            user.save()

            # Save the student instance (which is already linked to the user)
            student = form.save(commit=False)
            student.activate = request.POST.get('activate') == 'on' # Handle activate checkbox
            student.save()

            messages.success(request, "تم تحديث بيانات الطالب بنجاح!")
            return redirect('manage_students') # Redirect back to the manage students page
        else:
            messages.error(request, "حدث خطأ أثناء تحديث بيانات الطالب. يرجى مراجعة الأخطاء أدناه.")
            # Render the form with errors
            return render(request, 'admin/edit_student.html', {'form': form, 'student': student})
    else:
        # Populate the form with the existing student and user data
        # Need to pass initial data for User fields to the StudentForm
        initial_data = {
            'first_name': user.first_name,
            'last_name': user.last_name,
            'email': user.email,
            'username': user.username, # Display username but it won't be editable by default
            # Add other fields from the Student model
            'birthday': student.birthday,
            'nationality': student.nationality,
            'gender': student.gender,
            'language': student.language,
            'mobile': student.mobile,
            'whatsapp': student.whatsapp,
            'halaqa': student.halaqa,
            'activate': student.activate,
        }
        form = StudentForm(instance=student, initial=initial_data)

    return render(request, 'admin/edit_student.html', {'form': form, 'student': student})


@login_required
def delete_student_view(request, student_id):
    # Check if the user is an admin
    if not request.user.groups.filter(name='admin').exists():
        messages.error(request, "You do not have permission to access this page.")
        return JsonResponse({'message': 'Unauthorized'}, status=403)

    if request.method == 'POST':
        try:
            student = get_object_or_404(Students, id=student_id)
            # Optionally delete the associated Django User as well
            user = student.user
            student.delete()
            if user:
                user.delete()
            return JsonResponse({'message': 'تم الحذف بنجاح!'})
        except Students.DoesNotExist:
            return JsonResponse({'message': 'الطالب غير موجود.'}, status=404)
        except Exception as e:
            return JsonResponse({'message': f'حدث خطأ أثناء الحذف: {e}'}, status=500)

    return JsonResponse({'message': 'Invalid request method'}, status=405)


@login_required
def delete_teacher_view(request, teacher_id):
    # Check if the user is an admin
    if not request.user.groups.filter(name='admin').exists():
        messages.error(request, "You do not have permission to access this page.")
        return JsonResponse({'message': 'Unauthorized'}, status=403)

    if request.method == 'POST':
        try:
            teacher = get_object_or_404(Teachers, id=teacher_id)
            # Optionally delete the associated Django User as well
            user = teacher.user
            teacher.delete()
            if user:
                user.delete()
            return JsonResponse({'message': 'تم الحذف بنجاح!'})
        except Teachers.DoesNotExist:
            return JsonResponse({'message': 'المدرس غير موجود.'}, status=404)
        except Exception as e:
            return JsonResponse({'message': f'حدث خطأ أثناء الحذف: {e}'}, status=500)

    return JsonResponse({'message': 'Invalid request method'}, status=405)


@login_required
def add_typeread_view(request):
    if not request.user.groups.filter(name='admin').exists():
        messages.error(request, "You do not have permission to access this page.")
        return redirect('index')

    if request.method == 'POST':
        form = TypeReadForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "تم إضافة القراءة بنجاح!")
            return redirect('manage_typeread')
    else:
        form = TypeReadForm()

    return render(request, 'manage/add_typeread.html', {'form': form})

@login_required
def edit_typeread_view(request, typeread_id):
    if not request.user.groups.filter(name='admin').exists():
        messages.error(request, "You do not have permission to access this page.")
        return redirect('index')

    typeread = get_object_or_404(TypeRead, id=typeread_id)

    if request.method == 'POST':
        form = TypeReadForm(request.POST, instance=typeread)
        if form.is_valid():
            form.save()
            messages.success(request, "تم تحديث القراءة بنجاح!")
            return redirect('manage_typeread')
    else:
        form = TypeReadForm(instance=typeread)

    return render(request, 'manage/edit_typeread.html', {'form': form, 'typeread': typeread})

@login_required
def delete_typeread_view(request, typeread_id):
    if not request.user.groups.filter(name='admin').exists():
        messages.error(request, "You do not have permission to access this page.")
        return JsonResponse({'message': 'Unauthorized'}, status=403)

    if request.method == 'POST':
        typeread = get_object_or_404(TypeRead, id=typeread_id)
        try:
            typeread.delete()
            return JsonResponse({'message': 'تم الحذف بنجاح!'})
        except Exception as e:
            return JsonResponse({'message': f'حدث خطأ أثناء الحذف: {e}'}, status=500)

    return JsonResponse({'message': 'Invalid request method'}, status=405)


@login_required
def add_halaqa_view(request):
    if not request.user.groups.filter(name='admin').exists():
        messages.error(request, "You do not have permission to access this page.")
        return redirect('index')

    if request.method == 'POST':
        form = HalaqaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "تم إضافة الحلقة بنجاح!")
            return redirect('manage_halaqat')
    else:
        form = HalaqaForm()

    return render(request, 'manage/add_halaqa.html', {'form': form})

@login_required
def edit_halaqa_view(request, halaqa_id):
    if not request.user.groups.filter(name='admin').exists():
        messages.error(request, "You do not have permission to access this page.")
        return redirect('index')

    halaqa = get_object_or_404(Halaqa, id=halaqa_id)

    if request.method == 'POST':
        form = HalaqaForm(request.POST, instance=halaqa)
        if form.is_valid():
            form.save()
            messages.success(request, "تم تحديث الحلقة بنجاح!")
            return redirect('manage_halaqat')
    else:
        form = HalaqaForm(instance=halaqa)

    return render(request, 'manage/edit_halaqa.html', {'form': form, 'halaqa': halaqa})

@login_required
def add_msar_view(request):
    if not request.user.groups.filter(name='admin').exists():
        messages.error(request, "You do not have permission to access this page.")
        return redirect('index')

    if request.method == 'POST':
        form = MsarForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "تم إضافة المسار بنجاح!")
            return redirect('manage_msars')
    else:
        form = MsarForm()

    return render(request, 'manage/add_msar.html', {'form': form})
@login_required
def edit_msar_view(request, msar_id):
    if not request.user.groups.filter(name='admin').exists():
        messages.error(request, "You do not have permission to access this page.")
        return redirect('index')

    msar = get_object_or_404(Msar, id=msar_id)

    if request.method == 'POST':
        form = MsarForm(request.POST, instance=msar)
        if form.is_valid():
            form.save()
            messages.success(request, "تم تحديث المسار بنجاح!")
            return redirect('manage_msars')
    else:
        form = MsarForm(instance=msar)

    return render(request, 'manage/edit_msar.html', {'form': form, 'msar': msar})

def add_department(request):
    if not request.user.groups.filter(name='admin').exists():
        messages.error(request, "You do not have permission to access this page.")
        return redirect('index')

    if request.method == 'POST':
        form = DepartmentForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "تم إضافة القسم بنجاح!")
            return redirect('manage_sections')
    else:
        form = DepartmentForm()

    return render(request, 'manage/add_department.html', {'form': form})


@login_required
def edit_department(request, department_id):
    if not request.user.groups.filter(name='admin').exists():
        messages.error(request, "You do not have permission to access this page.")
        return redirect('index')

    department = get_object_or_404(Department, id=department_id)

    if request.method == 'POST':
        form = DepartmentForm(request.POST, instance=department)
        if form.is_valid():
            form.save()
            messages.success(request, "تم تحديث القسم بنجاح!")
            return redirect('manage_sections')
    else:
        form = DepartmentForm(instance=department)

    return render(request, 'manage/edit_department.html', {'form': form, 'department': department})


@login_required
def delete_department(request, department_id):
    if not request.user.groups.filter(name='admin').exists():
        messages.error(request, "You do not have permission to access this page.")
        return JsonResponse({'message': 'Unauthorized'}, status=403)

    if request.method == 'POST':
        department = get_object_or_404(Department, id=department_id)
        try:
            department.delete()
            return JsonResponse({'message': 'تم الحذف بنجاح!'})
        except Exception as e:
            return JsonResponse({'message': f'حدث خطأ أثناء الحذف: {e}'}, status=500)
    
    return JsonResponse({'message': 'Invalid request method'}, status=405)

@login_required
def delete_msar_view(request, msar_id):
    if not request.user.groups.filter(name='admin').exists():
        messages.error(request, "You do not have permission to access this page.")
        return JsonResponse({'message': 'Unauthorized'}, status=403)

    if request.method == 'POST':
        msar = get_object_or_404(Msar, id=msar_id)
        try:
            msar.delete()
            return JsonResponse({'message': 'تم الحذف بنجاح!'})
        except Exception as e:
            return JsonResponse({'message': f'حدث خطأ أثناء الحذف: {e}'}, status=500)

    return JsonResponse({'message': 'Invalid request method'}, status=405)

@login_required
def delete_halaqa_view(request, halaqa_id):
    if not request.user.groups.filter(name='admin').exists():
        messages.error(request, "You do not have permission to access this page.")
        return JsonResponse({'message': 'Unauthorized'}, status=403)

    if request.method == 'POST':
        halaqa = get_object_or_404(Halaqa, id=halaqa_id)
        try:
            halaqa.delete()
            return JsonResponse({'message': 'تم الحذف بنجاح!'})
        except Exception as e:
            return JsonResponse({'message': f'حدث خطأ أثناء الحذف: {e}'}, status=500)

    return JsonResponse({'message': 'Invalid request method'}, status=405)


@login_required
def superuser_dashboard(request):
    return render(request, 'teacher/superuser_dashboard.html')


def home(request):
    total_students = Students.objects.count() # Calculate total number of students
    return render(request, 'index.html', {'user': request.user})




def index(request):
    total_students = Students.objects.count()
    total_halaqat = Halaqa.objects.count()
    total_memorized_ayat = DailyMemorizationTask.objects.aggregate(Sum('memorized_ayat_count'))['memorized_ayat_count__sum'] or 0
    total_memorized_pages = DailyMemorizationTask.objects.aggregate(Sum('pages'))['pages__sum'] or 0
    total_memorized_partial_pages = DailyMemorizationTask.objects.aggregate(Sum('partial_page'))['partial_page__sum'] or 0
    total_pages_and_parts = total_memorized_pages + total_memorized_partial_pages
    context = {
        'total_students': total_students,
        'total_halaqat': total_halaqat,
        'total_memorized_ayat': total_memorized_ayat,
        'total_pages_and_parts': total_pages_and_parts
    }
    return render(request, 'index.html', context)




def halaqa_detail(request, id):
    halaqa = get_object_or_404(Halaqa, pk=id)
    students = Students.objects.filter(halaqa=halaqa)
    # تمرير اسم المستخدم في السياق
    username = request.user.username if request.user.is_authenticated else None
    return render(request, 'teacher/halaqa_detail.html', {'halaqa': halaqa, 'students': students, 'username': username})






from django.db.models import Sum

def memorization_history_view(request, halaqa_id):
    halaqa = get_object_or_404(Halaqa, pk=halaqa_id)
    # Get all students in the halaqa
    students_in_halaqa = Students.objects.filter(halaqa=halaqa)
    # Get all memorization tasks for students in this halaqa
    memorization_records = DailyMemorizationTask.objects.filter(student__in=students_in_halaqa).order_by('date')

    # Calculate total verses and pages/parts
    total_verses = memorization_records.aggregate(Sum('memorized_ayat_count'))['memorized_ayat_count__sum'] or 0
    total_pages = memorization_records.aggregate(Sum('pages'))['pages__sum'] or 0
    total_partial_pages = memorization_records.aggregate(Sum('partial_page'))['partial_page__sum'] or 0
    total_pages_parts = total_pages + total_partial_pages

    context = {
        'halaqa': halaqa,
        'memorization_records': memorization_records,
        'total_verses': total_verses,
        'total_pages_parts': total_pages_parts,
    }
    return render(request, 'teacher/memorization_history.html', context)


def student_detail(request, id):
    student = get_object_or_404(Students, id=id)
    return render(request, 'students/student_detail.html', {'student': student})




def about(request):
    return render(request, 'students/about.html')

def almaqraa(request):
    return render(request, 'almaqraa.html')

def teluse(request):
    return render(request, 'telus.html')






def Daily_Memorization_Task_select_student(request, halaqa_id):
    halaqa = get_object_or_404(Halaqa, pk=halaqa_id)
    students = Students.objects.filter(halaqa=halaqa)
    context = {
        'students': students,
        'halaqa': halaqa
    }
    return render(request, 'students/Daily_Memorization_Task_select_student.html', context)








#def Daily_Memorization_Task_add(request, student_id):
   # student = get_object_or_404(Students, id=student_id)
  #  if request.method == 'POST':
  #      form = DailyMemorizationTaskForm(request.POST)
    #    if form.is_valid():
   #         task = form.save(commit=False)
      #      task.student = student
      #      task.save()
     #       return redirect('success_page')  # يمكنك تحديد صفحة النجاح المناسبة هنا
    #else:
    #    form = DailyMemorizationTaskForm()
    
   #return render(request, 'students/Daily_Memorization_Task_add.html', {'form': form, 'student': student})




#from django.shortcuts import render, redirect
#from .forms import DailyMemorizationTaskForm

#def add_memorization_task(request):
#    if request.method == 'POST':
#        form = DailyMemorizationTaskForm(request.POST)
#        if form.is_valid():
#            form.save()
#            return redirect('success_url')
#    else:
#        form = DailyMemorizationTaskForm()
#
#    return render(request, 'students/Daily_Memorization_Task_add.html', {'form': form})






def success_page(request):
    return render(request, 'students/success.html')




# الفئات والسور المرتبطة بها
category_surahs = {
    "المبتدئين": ["الفاتحة", "الناس", "الفلق", "الإخلاص", "المسد", "النصر", "الكافرون", "الكوثر", "الماعون", "قريش", "الفيل", "الهمزة", "العصر", "التكاثر", "القارعة", "العاديات", "الزلزلة", "البينة", "القدر", "العلق", "التين", "الشرح", "الضحى", "الليل", "الشمس", "البلد", "الفجر", "الغاشية", "الأعلى", "الطارق", "البروج", "الإنشقاق", "المطففين", "الانفطار", "التكوير", "عبس", "النازعات", "النبأ"],
    "ثلاثة أجزاء": ["المرسلات", "الإنسان", "القيامة", "المدثر", "المزمل", "الجن", "نوح", "المعارج", "الحاقة", "القلم", "الملك", "التحريم", "الطلاق", "التغابن", "المنافقون", "الجمعة", "الصف", "الممتحنة", "الحشر", "المجادلة"],
    "خمسة أجزاء": ["الحديد", "الواقعة", "الرحمن", "القمر", "النجم", "الطور", "الذاريات", "ق", "الحجرات", "الفتح", "محمد", "الأحقاف"],
    "عشرة أجزاء": ["الجاثية", "الدخان", "الزخرف", "الشورى", "فصلت", "غافر", "الزمر", "ص", "الصافات", "يس", "فاطر", "سبأ", "الأحزاب", "السجدة", "لقمان", "الروم"],
    "خمسة عشر جزء": ["العنكبوت", "القصص", "النمل", "الشعراء", "الفرقان", "النور", "المؤمنون", "الحج", "الأنبياء", "طه", "مريم"],
    "عشرون جزء": ["الكهف", "الإسراء", "النحل", "الحجر", "إبراهيم", "الرعد", "يوسف", "هود", "يونس"],
    "خمسة وعشرون جزء": ["التوبة", "الأنفال", "الأعراف", "الأنعام", "المائدة", "النساء", "آل عمران", "البقرة"]
}
# عدد الآيات لكل سورة (بيانات تقريبية، تحتاج إلى التحقق الدقيق)
ayatCounts = {
    "الفاتحة": 7,
    "الناس": 6,
    "الفلق": 5,
    "الإخلاص": 4,
    "المسد": 5,
    "النصر": 3,
    "الكافرون": 6,
    "الكوثر": 3,
    "الماعون": 7,
    "قريش": 4,
    "الفيل": 5,
    "الهمزة": 9,
    "العصر": 3,
    "التكاثر": 8,
    "القارعة": 11,
    "العاديات": 11,
    "الزلزلة": 8,
    "البينة": 8,
    "القدر": 5,
    "العلق": 19,
    "التين": 8,
    "الشرح": 8,
    "الضحى": 11,
    "الليل": 21,
    "الشمس": 15,
    "البلد": 20,
    "الفجر": 30,
    "الغاشية": 26,
    "الأعلى": 19,
    "الطارق": 17,
    "البروج": 22,
    "الإنشقاق": 25,
    "المطففين": 36,
    "الانفطار": 19,
    "التكوير": 29,
    "عبس": 42,
    "النازعات": 46,
    "النبأ": 40,
    "المرسلات": 50,
    "الإنسان": 31,
    "القيامة": 40,
    "المدثر": 56,
    "المزمل": 20,
    "الجن": 28,
    "نوح": 28,
    "المعارج": 44,
    "الحاقة": 52,
    "القلم": 52,
    "الملك": 30,
    "التحريم": 12,
    "الطلاق": 12,
    "التغابن": 18,
    "المنافقون": 11,
    "الجمعة": 11,
    "الصف": 14,
    "الممتحنة": 13,
    "الحشر": 24,
    "المجادلة": 22,
    "الحديد": 29,
    "الواقعة": 96,
    "الرحمن": 78,
    "القمر": 55,
    "النجم": 62,
    "الطور": 49,
    "الذاريات": 60,
    "ق": 45,
    "الحجرات": 18,
    "الفتح": 29,
    "محمد": 38,
    "الأحقاف": 35,
    "الجاثية": 37,
    "الدخان": 59,
    "الزخرف": 89,
    "الشورى": 53,
    "فصلت": 54,
    "غافر": 85,
    "الزمر": 75,
    "ص": 88,
    "الصافات": 182,
    "يس": 83,
    "فاطر": 45,
    "سبأ": 54,
    "الأحزاب": 73,
    "السجدة": 30,
    "لقمان": 34,
    "الروم": 60,
    "العنكبوت": 69,
    "القصص": 88,
    "النمل": 93,
    "الشعراء": 227,
    "الفرقان": 77,
    "النور": 64,
    "المؤمنون": 118,
    "الحج": 78,
    "الأنبياء": 112,
    "طه": 135,
    "مريم": 98,
    "الكهف": 110,
    "الإسراء": 111,
    "النحل": 128,
    "الحجر": 99,
    "إبراهيم": 52,
    "الرعد": 43,
    "يوسف": 111,
    "هود": 123,
    "يونس": 109,
    "التوبة": 129,
    "الأنفال": 75,
    "الأعراف": 206,
    "الأنعام": 165,
    "المائدة": 120,
    "النساء": 176,
    "آل عمران": 200,
    "البقرة": 286,
}

def Daily_Memorization_Task_add(request, student_id):
    student = get_object_or_404(Students, id=student_id)
    if request.method == 'POST':
        form = DailyMemorizationTaskForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('success_page', student_id=student_id)
    else:
        form = DailyMemorizationTaskForm(initial={'student': student})
    return render(request, 'students/Daily_Memorization_Task_add.html', {'form': form, 'student': student})

def save_memorization_task(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            print("Received Data:", data)  # Debug print
            # Validate required fields are present
            required_fields = ["student_id", "MemorizationTask", "previous_surah", "previous_ayat", "current_surah", "current_ayat", "pages", "partial_page", "grade", "memorized_ayat_count"]
            for field in required_fields:
                if field not in data:
                    raise KeyError(f"Missing required field: {field}")

            # Fetch the student object
            student = get_object_or_404(Students, id=data["student_id"])

            task = DailyMemorizationTask(
                student=student, # Assign the student object
                MemorizationTask=data["MemorizationTask"],
                previous_surah=data["previous_surah"],
                previous_ayat=data["previous_ayat"],
                current_surah=data["current_surah"],
                current_ayat=data["current_ayat"],
                pages=data["pages"],
                partial_page=data["partial_page"],
                grade=data["grade"],
                memorized_ayat_count=data["memorized_ayat_count"]
            )
            task.save()
            # Include student_id and halaqa_id in the success response
            return JsonResponse({"message": "تم حفظ المهمة بنجاح!", "student_id": student.id, "halaqa_id": student.halaqa.id if student.halaqa else None})
        except (KeyError, ValueError) as e:
            print("Error:", e)  # Debug print
            return JsonResponse({"message": "فشل في حفظ المهمة.", "error": str(e)}, status=400)
        except Students.DoesNotExist:
             print(f"Error: Student with id {data.get('student_id')} not found.") # Debug print
             return JsonResponse({"message": "فشل في حفظ المهمة.", "error": "Student not found."}, status=404)
        except Exception as e:
            print("Unexpected Error:", e) # Debug print
            return JsonResponse({"message": "فشل في حفظ المهمة.", "error": str(e)}, status=500)
    return JsonResponse({"message": "فشل في حفظ المهمة."}, status=405) # Method Not Allowed

def get_surahs_by_category(request, category):
    surahs = category_surahs.get(category, [])
    return JsonResponse(surahs, safe=False)

def success_page(request, student_id):
    student = get_object_or_404(Students, id=student_id)
    return render(request, 'students/success.html', {'student': student})





def Daily_Review_Task_add(request, student_id):
    student = get_object_or_404(Students, id=student_id)
    if request.method == 'POST':
        form = DailyReviewTaskForm(request.POST)
        if form.is_valid():
            daily_task = form.save(commit=False)
            daily_task.student = student
            daily_task.save()
            return redirect('success_review', student_id=student_id)
    else:
        form = DailyReviewTaskForm(initial={'student': student})
    return render(request, 'students/Daily_Review_Task_add.html', {'form': form, 'student': student})

def save_daily_review_task(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            print("Received Data:", data)  # طباعة البيانات للتحقق منها
            student = get_object_or_404(Students, id=data["student_id"])
            task = DailyReviewTask(
                student=student,
                ReviewTask=data["ReviewTask"],  # تأكد من مطابقة الاسماء 
                previous_surah=data["previous_surah"],
                previous_ayat=data["previous_ayat"],
                current_surah=data["current_surah"],
                current_ayat=data["current_ayat"],
                pages=data["pages"],
                partial_page=data["partial_page"],
                grade=data["grade"],
                reviewed_ayat_count=data["reviewed_ayat_count"]
            )
            task.save()
            # Include student_id and halaqa_id in the success response
            return JsonResponse({"message": "تم حفظ المهمة بنجاح!", "student_id": student.id, "halaqa_id": student.halaqa.id if student.halaqa else None})
        except (KeyError, ValueError) as e:
            print("Error:", e)  # طباعة الخطأ للتحقق منه
            return JsonResponse({"message": "فشل في حفظ المهمة.", "error": str(e)}, status=400)
        except Students.DoesNotExist:
             print(f"Error: Student with id {data.get('student_id')} not found.") # Debug print
             return JsonResponse({"message": "فشل في حفظ المهمة.", "error": "Student not found."}, status=404)
        except Exception as e:
            print("Unexpected Error:", e) # Debug print
            return JsonResponse({"message": "فشل في حفظ المهمة.", "error": str(e)}, status=500)
    return JsonResponse({"message": "فشل في حفظ المهمة."}, status=405) # Method Not Allowed

def get_surahs_by_category(request):
    category = request.GET.get('category')
    surahs = category_surahs.get(category, [])
    return JsonResponse(surahs, safe=False)

def success_review(request, student_id):
    student = get_object_or_404(Students, id=student_id)
    return render(request, 'students/success_review.html', {'student': student})




def Daily_Review_Task_select_student(request, halaqa_id):
    students = Students.objects.all()
    halaqa = get_object_or_404(Halaqa, pk=halaqa_id)
    context = {
        'students': students,
        'halaqa': halaqa
    }
    return render(request, 'students/Daily_Review_Task_select_student.html', context)






category_surahs = {
    "المبتدئين": ["الفاتحة", "الناس", "الفلق", "الإخلاص", "المسد", "النصر", "الكافرون", "الكوثر", "الماعون", "قريش", "الفيل", "الهمزة", "العصر", "التكاثر", "القارعة", "العاديات", "الزلزلة", "البينة", "القدر", "العلق", "التين", "الشرح", "الضحى", "الليل", "الشمس", "البلد", "الفجر", "الغاشية", "الأعلى", "الطارق", "البروج", "الإنشقاق", "المطففين", "الانفطار", "التكوير", "عبس", "النازعات", "النبأ"],
    "ثلاثة أجزاء": ["المرسلات", "الإنسان", "القيامة", "المدثر", "المزمل", "الجن", "نوح", "المعارج", "الحاقة", "القلم", "الملك", "التحريم", "الطلاق", "التغابن", "المنافقون", "الجمعة", "الصف", "الممتحنة", "الحشر", "المجادلة"],
    "خمسة أجزاء": ["الحديد", "الواقعة", "الرحمن", "القمر", "النجم", "الطور", "الذاريات", "ق", "الحجرات", "الفتح", "محمد", "الأحقاف"],
    "عشرة أجزاء": ["الجاثية", "الدخان", "الزخرف", "الشورى", "فصلت", "غافر", "الزمر", "ص", "الصافات", "يس", "فاطر", "سبأ", "الأحزاب", "السجدة", "لقمان", "الروم"],
    "خمسة عشر جزء": ["العنكبوت", "القصص", "النمل", "الشعراء", "الفرقان", "النور", "المؤمنون", "الحج", "الأنبياء", "طه", "مريم"],
    "عشرون جزء": ["الكهف", "الإسراء", "النحل", "الحجر", "إبراهيم", "الرعد", "يوسف", "هود", "يونس"],
    "خمسة وعشرون جزء": ["التوبة", "الأنفال", "الأعراف", "الأنعام", "المائدة", "النساء", "آل عمران", "البقرة"]
}

def Daily_Talqeen_Task_add(request, student_id):
    student = get_object_or_404(Students, id=student_id)
    if request.method == 'POST':
        form = DailyTalqeenTaskForm(request.POST)
        if form.is_valid():
            daily_task = form.save(commit=False)
            daily_task.student = student
            daily_task.save()
            return redirect('success_talqeen', student_id=student_id)
    else:
        form = DailyTalqeenTaskForm(initial={'student': student})
    return render(request, 'students/Daily_Talqeen_Task_add.html', {'form': form, 'student': student})

def save_daily_talqeen_task(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            print("Received Data:", data)  # طباعة البيانات للتحقق منها
            student = get_object_or_404(Students, id=data["student_id"])
            task = DailyTalqeenTask(
                student=student,
                TalqeenTask=data["TalqeenTask"],  # تأكد من مطابقة الأسماء
                previous_surah=data["previous_surah"],
                previous_ayat=data["previous_ayat"],
                current_surah=data["current_surah"],
                current_ayat=data["current_ayat"],
                pages=data["pages"],
                partial_page=data["partial_page"],
                grade=data["grade"],
                Talqeen_ayat_count=data["Talqeen_ayat_count"]
            )
            task.save()
            # Include student_id and halaqa_id in the success response
            return JsonResponse({"message": "تم حفظ المهمة بنجاح!", "student_id": student.id, "halaqa_id": student.halaqa.id if student.halaqa else None})
        except (KeyError, ValueError) as e:
            print("Error:", e)  # طباعة الخطأ للتحقق منه
            return JsonResponse({"message": "فشل في حفظ المهمة.", "error": str(e)}, status=400)
        except Students.DoesNotExist:
             print(f"Error: Student with id {data.get('student_id')} not found.") # Debug print
             return JsonResponse({"message": "فشل في حفظ المهمة.", "error": "Student not found."}, status=404)
        except Exception as e:
            print("Unexpected Error:", e) # Debug print
            return JsonResponse({"message": "فشل في حفظ المهمة.", "error": str(e)}, status=500)
    return JsonResponse({"message": "فشل في حفظ المهمة."}, status=405) # Method Not Allowed

# Note: There are multiple definitions of this function in this file.
# This is the last definition and is used by the '/get_surahs/' URL pattern.
def get_surahs_by_category(request):
    category = request.GET.get('category')
    surahs = category_surahs.get(category, [])
    return JsonResponse(surahs, safe=False)




def get_ayahs_by_surah(request):
    surah = request.GET.get('surah')
    ayah_count = ayatCounts.get(surah, 0) # Use the global ayatCounts
    return JsonResponse({'ayah_count': ayah_count})

def Daily_Talqeen_Task_select_student(request, halaqa_id):
    students = Students.objects.all()
    halaqa = get_object_or_404(Halaqa, pk=halaqa_id)
    context = {
        'students': students,
        'halaqa': halaqa
    }
    return render(request, 'students/Daily_Talqeen_Task_select_student.html', context)


def success_talqeen(request, student_id):
    student = get_object_or_404(Students, id=student_id)
    return render(request, 'students/success_talqeen.html', {'student': student})




def student_detail_view(request, student_id):
    student = get_object_or_404(Students, id=student_id)
    last_entry = DailyMemorizationTask.objects.filter(student=student).latest('id')
    last_review_entry = DailyReviewTask.objects.filter(student=student).latest('id')
    last_talqeen_entry = DailyTalqeenTask.objects.filter(student=student).latest('id')
    weekly_plan = WeeklyPlan.objects.filter(student=student).latest('id')  # الحصول على آخر خطة أسبوعية
    context = {
        'student': student,
        'last_entry': last_entry,
        'last_review_entry': last_review_entry,
        'last_talqeen_entry': last_talqeen_entry,
        'weekly_plan': weekly_plan,
    }
    return render(request, 'students/student_detail.html', context)

def student_home_view(request, student_id):
    student = get_object_or_404(Students, id=student_id)
    last_entry = DailyMemorizationTask.objects.filter(student=student).latest('id')
    last_review_entry = DailyReviewTask.objects.filter(student=student).latest('id')
    last_talqeen_entry = DailyTalqeenTask.objects.filter(student=student).latest('id')
    weekly_plan = WeeklyPlan.objects.filter(student=student).latest('id')  # الحصول على آخر خطة أسبوعية
    context = {
        'student': student,
        'last_entry': last_entry,
        'last_review_entry': last_review_entry,
        'last_talqeen_entry': last_talqeen_entry,
        'weekly_plan': weekly_plan,
        
        
    }
    return render(request, 'students/student_home.html', context)






def weekly_plan_view(request, student_id):
    student = get_object_or_404(Students, id=student_id)
    # Get the existing weekly plan for the student, if it exists
    try:
        weekly_plan = WeeklyPlan.objects.get(student=student)
    except WeeklyPlan.DoesNotExist:
        weekly_plan = None

    if request.method == 'POST':
        # If a plan exists, update that instance
        form = WeeklyPlanForm(request.POST, instance=weekly_plan)
        if form.is_valid():
            plan = form.save(commit=False)
            plan.student = student
            plan.save()
            # إعادة التوجيه إلى صفحة التفاصيل بعد الحفظ
            return redirect('student_detail', student_id=student.id)
    else:
        # If a plan exists, initialize the form with that instance
        if weekly_plan:
            form = WeeklyPlanForm(instance=weekly_plan)
        else:
            # Otherwise, initialize an empty form
            form = WeeklyPlanForm()

    context = {
        'student': student,
        'form': form,
    }
    return render(request, 'students/weekly_plan.html', context)




def weekly_plan_select_student(request, halaqa_id):
    students = Students.objects.filter(halaqa=halaqa_id)
    halaqa = get_object_or_404(Halaqa, pk=halaqa_id)
    context = {
        'students': students,
        'halaqa': halaqa
    }
    return render(request, 'students/weekly_plan_select_student.html', context)




def weekly_lecture_list(request):
    lectures = WeeklyLecture.objects.all()
    return render(request, 'weekly_lectures/lecture_list.html', {'lectures': lectures})

def weekly_lecture_detail(request, lecture_id):
    lecture = get_object_or_404(WeeklyLecture, id=lecture_id)
    return render(request, 'weekly_lectures/lecture_detail.html', {'lecture': lecture})

def add_weekly_lecture(request):
    if request.method == 'POST':
        form = WeeklyLectureForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('lecture_list')
    else:
        form = WeeklyLectureForm()
    return render(request, 'weekly_lectures/add_lecture.html', {'form': form})

@login_required
def delete_weekly_lecture_view(request, lecture_id):
    print(f"delete_weekly_lecture_view called for lecture_id: {lecture_id}") # Log view entry
    # Check if the user is an admin
    if not request.user.groups.filter(name='admin').exists():
        messages.error(request, "You do not have permission to access this page.")
        print("Unauthorized access attempt to delete_weekly_lecture_view") # Log unauthorized access
        return JsonResponse({'message': 'Unauthorized'}, status=403)

    if request.method == 'POST':
        try:
            print(f"Attempting to delete lecture with ID: {lecture_id}") # Log deletion attempt
            lecture = get_object_or_404(WeeklyLecture, id=lecture_id)
            lecture.delete()
            print(f"Lecture with ID {lecture_id} deleted successfully.") # Log successful deletion
            return JsonResponse({'message': 'تم الحذف بنجاح!'})
        except WeeklyLecture.DoesNotExist:
            print(f"Lecture with ID {lecture_id} not found.") # Log lecture not found
            return JsonResponse({'message': 'المحاضرة غير موجودة.'}, status=404)
        except Exception as e:
            print(f"Error deleting lecture with ID {lecture_id}: {e}") # Log deletion error
            return JsonResponse({'message': f'حدث خطأ أثناء الحذف: {e}'}, status=500)

    print("Invalid request method for delete_weekly_lecture_view") # Log invalid method
    return JsonResponse({'message': 'Invalid request method'}, status=405)




def latest_lecture_view(request):
    # جلب المحاضرة الأسبوعية الأخيرة
    lecture = WeeklyLecture.objects.order_by('-created_at').first()
    context = {
        'lecture': lecture,
    }
    return render(request, 'weekly_lectures/latest_lecture.html', context)



# whatsapp_notifier.py




def send_whatsapp_message():
    # إعداد WebDriver
    driver = webdriver.Chrome(executable_path='/path/to/chromedriver')
    driver.get('https://web.whatsapp.com')

    # انتظر حتى يقوم المستخدم بمسح كود الـ QR
    input('Press Enter after scanning QR code')

    # ابحث عن الشخص أو المجموعة باستخدام اسمهم
    search_box = driver.find_element_by_xpath('//div[@contenteditable="true"][@data-tab="3"]')
    search_box.send_keys('اسم الشخص أو المجموعة')
    search_box.send_keys(Keys.RETURN)

    # اكتب وأرسل الرسالة
    message_box = driver.find_element_by_xpath('//div[@contenteditable="true"][@data-tab="6"]')
    message_box.send_keys('هذه رسالة اختبار عبر WhatsApp')
    message_box.send_keys(Keys.RETURN)

    time.sleep(5)  # انتظر قليلاً قبل إغلاق المتصفح
    driver.quit()

# views.py

def notify_user(request):
    send_whatsapp_message()
    return render(request, 'weekly_lectures/notification_sent.html')




from django.shortcuts import render, redirect
from .models import Message, Halaqa, Students, Teachers
from .forms import MessageForm
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.db.models import Q

@login_required
def chat_view(request, username):
    # احصل على الطالب بناءً على اسم المستخدم
    student = Students.objects.get(username=username)
    # احصل على كائن المستخدم المرتبط بالطالب
    student_user = User.objects.get(username=student.username)
    # احصل على الأستاذ المرتبط بالحلقة
    teacher = student.halaqa.teacher

    if not teacher or not student:
        return redirect('home')

    messages = Message.objects.filter(
        Q(sender=request.user, recipient=teacher.user) |
        Q(sender=teacher.user, recipient=request.user) |
        Q(sender=request.user, recipient=student_user) |
        Q(sender=student_user, recipient=request.user)
    ).order_by('created_at')

    if request.method == 'POST':
        form = MessageForm(request.POST)
        if form.is_valid():
            message = form.save(commit=False)
            message.sender = request.user
            if request.user == student_user:
                message.recipient = teacher.user
            else:
                message.recipient = student_user
            message.halaqa = student.halaqa
            message.save()
            return redirect('chat', username=username)
    else:
        form = MessageForm()

    return render(request, 'students/chat.html', {
        'messages': messages,
        'user': request.user,
        'form': form,
        'teacher': teacher,
        'student': student,
        'halaqa': student .halaqa
    })






from django.shortcuts import render, redirect
from .models import Message, Students, Teachers, Admin
from .forms import MessageForm
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.db.models import Q

@login_required
def chat_with_admin_view(request, username):
    # احصل على المستخدم الحالي
    current_user = request.user

    # تحقق مما إذا كان المستخدم الحالي هو طالب أو مدرس أو أدمن
    is_student = False
    is_teacher = False
    is_admin = False
    student = None
    teacher = None
    admin = None

    # تحقق مما إذا كان المستخدم الحالي هو طالب
    try:
        student = Students.objects.get(user=current_user)
        is_student = True
    except Students.DoesNotExist:
        pass

    # تحقق مما إذا كان المستخدم الحالي هو مدرس
    try:
        teacher = Teachers.objects.get(user=current_user)
        is_teacher = True
    except Teachers.DoesNotExist:
        pass

    # تحقق مما إذا كان المستخدم الحالي هو أدمن
    try:
        admin = Admin.objects.get(user=current_user)
        is_admin = True
    except Admin.DoesNotExist:
        pass

    # إذا كان المستخدم ليس طالبًا ولا مدرسًا ولا أدمن، أعد توجيهه إلى صفحة خطأ
    if not is_student and not is_teacher and not is_admin:
        return render(request, 'students/error.html', {'message': 'لم يتم العثور على كائن الطالب أو المدرس أو الأدمن. يرجى التحقق من بيانات المستخدم.'})

    # احصل على كائن المستخدم المرتبط بالأدمن بناءً على اسم المستخدم
    try:
        admin_user = User.objects.get(username=username)
        admin = Admin.objects.get(user=admin_user)
    except (User.DoesNotExist, Admin.DoesNotExist):
        return render(request, 'students/error.html', {'message': 'لم يتم العثور على كائن الأدمن المحدد. يرجى التحقق من اسم المستخدم.'})

    # اجلب الرسائل بين المستخدم الحالي والأدمن
    messages = Message.objects.filter(
        (Q(sender=current_user) & Q(receiver=admin_user)) |
        (Q(sender=admin_user) & Q(receiver=current_user))
    ).order_by('timestamp')

    if request.method == 'POST':
        form = MessageForm(request.POST)
        if form.is_valid():
            message = form.save(commit=False)
            message.sender = current_user
            message.receiver = admin_user
            message.save()
            return redirect('chat_with_admin', username=username)
    else:
        form = MessageForm()

    return render(request, 'students/chat_with_admin.html', {
        'messages': messages,
        'current_user': current_user,
        'form': form,
        'admin': admin,
        'student': student,
        'teacher': teacher
    })

@login_required
def assign_halaqa_to_student(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            student_id = data.get('student_id')
            halaqa_id = data.get('halaqa_id')

            if not student_id or not halaqa_id:
                return JsonResponse({'success': False, 'error': 'Missing student_id or halaqa_id'}, status=400)

            student = Students.objects.get(id=student_id)
            halaqa = Halaqa.objects.get(id=halaqa_id)

            student.halaqa = halaqa
            student.save()

            return JsonResponse({'success': True})

        except Students.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Student not found'}, status=404)
        except Halaqa.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Halaqa not found'}, status=404)
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    else:
        return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=405)





def student_detail_view(request, student_id):
    student = get_object_or_404(Students, id=student_id)
    try:
        last_entry = DailyMemorizationTask.objects.filter(student=student).latest('id')
    except DailyMemorizationTask.DoesNotExist:
        last_entry = None
    try:
        last_review_entry = DailyReviewTask.objects.filter(student=student).latest('id')
    except DailyReviewTask.DoesNotExist:
        last_review_entry = None
    try:
        last_talqeen_entry = DailyTalqeenTask.objects.filter(student=student).latest('id')
    except DailyTalqeenTask.DoesNotExist:
        last_talqeen_entry = None
    try:
        weekly_plan = WeeklyPlan.objects.filter(student=student).latest('id')  # الحصول على آخر خطة أسبوعية
    except WeeklyPlan.DoesNotExist:
        weekly_plan = None

    student_links = student.links.all() # Fetch all links for the student
    context = {
        'student': student,
        'last_entry': last_entry,
        'last_review_entry': last_review_entry,
        'last_talqeen_entry': last_talqeen_entry,
        'weekly_plan': weekly_plan,
        'student_links': student_links, # Add student_links to the context
    }
    return render(request, 'students/student_detail.html', context)

def student_home_view(request, student_id):
    student = get_object_or_404(Students, id=student_id)
    try:
        last_entry = DailyMemorizationTask.objects.filter(student=student).latest('id')
    except DailyMemorizationTask.DoesNotExist:
        last_entry = None
    try:
        last_review_entry = DailyReviewTask.objects.filter(student=student).latest('id')
    except DailyReviewTask.DoesNotExist:
        last_review_entry = None
    try:
        last_talqeen_entry = DailyTalqeenTask.objects.filter(student=student).latest('id')
    except DailyTalqeenTask.DoesNotExist:
        last_talqeen_entry = None
    try:
        weekly_plan = WeeklyPlan.objects.filter(student=student).latest('id')  # الحصول على آخر خطة أسبوعية
    except WeeklyPlan.DoesNotExist:
        weekly_plan = None

    context = {
        'student': student,
        'last_entry': last_entry,
        'last_review_entry': last_review_entry,
        'last_talqeen_entry': last_talqeen_entry,
        'weekly_plan': weekly_plan,


    }
    return render(request, 'students/student_home.html', context)

@login_required
def manage_links_view(request):
    # Check if the user is an admin
    if not request.user.groups.filter(name='admin').exists():
        messages.error(request, "You do not have permission to access this page.")
        return redirect('index') # Or a suitable unauthorized page

    if request.method == 'POST':
        form = StudentLinkForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "تم إضافة الرابط بنجاح!")
            return redirect('manage_links') # Redirect back to the same page
        else:
            messages.error(request, "حدث خطأ أثناء إضافة الرابط. يرجى مراجعة النموذج.")
    else:
        form = StudentLinkForm() # An empty form for adding new links

    # Fetch all existing links
    links = StudentLink.objects.all().select_related('student', 'student__halaqa')

    context = {
        'form': form,
        'links': links,
    }

    return render(request, 'admin/manage_links.html', context)